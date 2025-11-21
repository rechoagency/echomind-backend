"""
Reports Router - Excel export for Monday/Thursday content delivery
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date, timedelta
import pandas as pd
import io
import logging

from supabase_client import supabase
from services.knowledge_matchback_service import KnowledgeMatchbackService

logger = logging.getLogger(__name__)
router = APIRouter()
knowledge_service = KnowledgeMatchbackService(supabase)


@router.get("/reports/{client_id}/weekly-content")
async def get_weekly_content_report(
    client_id: str,
    delivery_batch: Optional[str] = None,
    format: str = "excel"
):
    """
    Generate Monday/Thursday content delivery report with profile assignments
    
    Args:
        client_id: Client UUID
        delivery_batch: Optional batch ID (e.g., 'MON-2025-W47')
        format: 'excel' or 'json'
        
    Returns:
        Excel file with:
        - Opportunity details
        - Suggested content
        - Assigned Reddit profile
        - Profile stats (karma, last posted)
        - Posting instructions
    """
    try:
        logger.info(f"📊 Generating content report for client {client_id}")
        
        # Get client data
        client_response = supabase.table('clients').select('company_name').eq('client_id', client_id).single().execute()
        if not client_response.data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        company_name = client_response.data['company_name']
        
        # Get generated content for this batch or recent
        query = supabase.table('content_delivered').select('*').eq('client_id', client_id)
        
        if delivery_batch:
            query = query.eq('delivery_batch', delivery_batch)
        else:
            # Get most recent batch (last 7 days)
            cutoff_date = (datetime.utcnow() - timedelta(days=7)).isoformat()
            query = query.gte('delivery_time', cutoff_date)
        
        content_response = query.order('delivery_time', desc=True).limit(50).execute()
        
        if not content_response.data or len(content_response.data) == 0:
            raise HTTPException(status_code=404, detail="No content found for this period")
        
        content_items = content_response.data
        
        logger.info(f"✅ Found {len(content_items)} content items")
        
        # Build Excel-friendly data
        excel_data = []
        
        for item in content_items:
            excel_data.append({
                'Opportunity ID': item.get('opportunity_id', 'N/A'),
                'Subreddit': f"r/{item.get('subreddit', 'unknown')}",
                'Content Type': item.get('content_type', 'reply').upper(),
                'Thread Title/Context': _get_thread_context(item.get('reddit_item_id')),
                'Suggested Response': item.get('content_text', ''),
                '📝 POST AS (Reddit Username)': item.get('profile_username', 'NO_PROFILE_ASSIGNED'),
                'Profile Karma': item.get('profile_karma', 'N/A'),
                'Brand Mentioned?': 'YES' if item.get('brand_mentioned') else 'NO',
                'Product Mentioned?': 'YES' if item.get('product_mentioned') else 'NO',
                'Word Count': item.get('word_count', 0),
                'Generated At': item.get('delivery_time', ''),
                'Delivery Batch': item.get('delivery_batch', 'N/A'),
                'Reddit Post URL': f"https://reddit.com{item.get('reddit_item_id', '')}" if item.get('reddit_item_id') else 'N/A'
            })
        
        if format == 'json':
            return {
                "success": True,
                "client_name": company_name,
                "content_count": len(excel_data),
                "content": excel_data
            }
        
        # Generate Excel file
        df = pd.DataFrame(excel_data)
        
        # Create Excel writer with formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Content Delivery', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Content Delivery']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)  # Cap at 100
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Highlight "POST AS" column
            from openpyxl.styles import PatternFill, Font
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
            
            # Find "POST AS" column
            for idx, col in enumerate(df.columns, 1):
                if 'POST AS' in col:
                    cell = worksheet.cell(row=1, column=idx)
                    cell.fill = yellow_fill
                    cell.font = bold_font
                    
                    # Highlight cells in this column too
                    for row in range(2, len(df) + 2):
                        worksheet.cell(row=row, column=idx).fill = PatternFill(
                            start_color="FFFFCC",
                            end_color="FFFFCC",
                            fill_type="solid"
                        )
        
        output.seek(0)
        
        # Generate filename
        today = date.today().strftime("%Y-%m-%d")
        filename = f"{company_name.replace(' ', '_')}_Content_Delivery_{today}.xlsx"
        
        logger.info(f"✅ Generated Excel report: {filename}")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error generating report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _get_thread_context(reddit_item_id: str) -> str:
    """
    Get thread title/context from opportunities table
    Fallback to Reddit ID if not found
    """
    try:
        if not reddit_item_id:
            return "N/A"
        
        # Query opportunities for this Reddit item
        response = supabase.table('opportunities') \
            .select('thread_title') \
            .eq('reddit_id', reddit_item_id) \
            .single() \
            .execute()
        
        if response.data:
            return response.data.get('thread_title', reddit_item_id)
        
        return reddit_item_id
    
    except:
        return reddit_item_id


@router.get("/reports/{client_id}/profile-analytics")
async def get_profile_analytics(client_id: str):
    """
    Get per-profile performance analytics
    
    Returns:
        Profile stats including:
        - Posts this week
        - Active subreddits
        - Karma level
        - Last post time
    """
    try:
        from datetime import timedelta
        
        # Get profiles
        profiles_response = supabase.table('client_reddit_profiles') \
            .select('*') \
            .eq('client_id', client_id) \
            .eq('is_active', True) \
            .execute()
        
        if not profiles_response.data:
            return {
                "success": False,
                "message": "No Reddit profiles configured"
            }
        
        profiles = profiles_response.data
        
        # Get content stats per profile (last 7 days)
        cutoff_date = (datetime.utcnow() - timedelta(days=7)).isoformat()
        
        content_response = supabase.table('content_delivered') \
            .select('profile_id, profile_username, subreddit, delivery_time') \
            .eq('client_id', client_id) \
            .gte('delivery_time', cutoff_date) \
            .execute()
        
        content_by_profile = {}
        if content_response.data:
            for item in content_response.data:
                profile_id = item.get('profile_id')
                if profile_id:
                    if profile_id not in content_by_profile:
                        content_by_profile[profile_id] = {
                            'posts': 0,
                            'subreddits': set(),
                            'last_posted': None
                        }
                    
                    content_by_profile[profile_id]['posts'] += 1
                    content_by_profile[profile_id]['subreddits'].add(item.get('subreddit'))
                    
                    delivery_time = item.get('delivery_time')
                    if delivery_time:
                        if not content_by_profile[profile_id]['last_posted'] or delivery_time > content_by_profile[profile_id]['last_posted']:
                            content_by_profile[profile_id]['last_posted'] = delivery_time
        
        # Build analytics
        analytics = []
        
        for profile in profiles:
            profile_id = profile.get('id')
            stats = content_by_profile.get(profile_id, {'posts': 0, 'subreddits': set(), 'last_posted': None})
            
            analytics.append({
                'username': profile.get('username'),
                'profile_type': profile.get('profile_type'),
                'current_karma': profile.get('current_karma', 0),
                'posts_this_week': stats['posts'],
                'active_subreddits': list(stats['subreddits']),
                'subreddit_count': len(stats['subreddits']),
                'last_posted': stats['last_posted'],
                'target_subreddits': profile.get('target_subreddits', []),
                'is_active': profile.get('is_active'),
                'created_at': profile.get('created_at')
            })
        
        return {
            "success": True,
            "client_id": client_id,
            "profile_count": len(analytics),
            "profiles": analytics
        }
    
    except Exception as e:
        logger.error(f"❌ Error getting profile analytics: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/{client_id}/knowledge-base-stats")
async def get_knowledge_base_stats(client_id: str):
    """
    Get statistics about client's knowledge base for thought leadership content
    
    Args:
        client_id: Client UUID
        
    Returns:
        Knowledge base statistics:
        - documents_uploaded: Number of documents in knowledge base
        - knowledge_chunks: Total number of embeddable insights
        - avg_chunks_per_document: Average granularity
        - estimated_coverage_kb: Approximate data volume
    """
    try:
        logger.info(f"📚 Getting knowledge base stats for client {client_id}")
        
        # Get knowledge base statistics
        stats = knowledge_service.get_knowledge_base_stats(client_id)
        
        # Get recent usage analytics
        # Check how many posts in the last week cited knowledge insights
        one_week_ago = (datetime.now() - timedelta(days=7)).isoformat()
        
        content_response = supabase.table('content_delivered') \
            .select('id, metadata') \
            .eq('client_id', client_id) \
            .gte('delivery_time', one_week_ago) \
            .execute()
        
        posts_with_insights = 0
        total_insights_cited = 0
        
        if content_response.data:
            for content in content_response.data:
                metadata = content.get('metadata', {})
                insights_used = metadata.get('knowledge_insights_count', 0)
                if insights_used > 0:
                    posts_with_insights += 1
                    total_insights_cited += insights_used
        
        total_posts = len(content_response.data) if content_response.data else 0
        usage_rate = round((posts_with_insights / total_posts * 100), 1) if total_posts > 0 else 0
        
        return {
            "success": True,
            "client_id": client_id,
            "knowledge_base": stats,
            "usage_last_7_days": {
                "total_posts": total_posts,
                "posts_with_insights": posts_with_insights,
                "total_insights_cited": total_insights_cited,
                "usage_rate_percentage": usage_rate,
                "avg_insights_per_post": round(total_insights_cited / total_posts, 2) if total_posts > 0 else 0
            }
        }
    
    except Exception as e:
        logger.error(f"❌ Error getting knowledge base stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))
