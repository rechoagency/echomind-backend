"""
Admin Router - Client Management Operations
Includes: Delete clients with confirmation, bulk operations
"""

from fastapi import APIRouter, HTTPException, BackgroundTasks
from pydantic import BaseModel
from typing import List
import logging
import os
from datetime import datetime
from openai import AsyncOpenAI

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/admin", tags=["Admin"])

supabase = None

def get_supabase():
    global supabase
    if supabase is None:
        from supabase_client import get_supabase_client
        supabase = get_supabase_client()
    return supabase


class DeleteClientRequest(BaseModel):
    client_id: str
    confirmation: bool = False  # Must be True to actually delete


@router.delete("/clients/{client_id}")
async def delete_client(client_id: str, confirmation: bool = False):
    """
    Delete a client and ALL associated data
    
    Requires confirmation=true to actually delete
    Without confirmation, returns what would be deleted
    """
    try:
        supabase = get_supabase()
        
        # Get client info first
        client = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        
        if not client.data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        client_data = client.data[0]
        
        # Count associated data
        opportunities = supabase.table("opportunities").select("opportunity_id", count="exact").eq("client_id", client_id).execute()
        documents = supabase.table("document_uploads").select("id", count="exact").eq("client_id", client_id).execute()
        calendars = supabase.table("content_calendars").select("id", count="exact").eq("client_id", client_id).execute()
        
        summary = {
            "client": client_data.get("company_name"),
            "client_id": client_id,
            "will_delete": {
                "opportunities": len(opportunities.data) if opportunities.data else 0,
                "documents": len(documents.data) if documents.data else 0,
                "calendars": len(calendars.data) if calendars.data else 0
            }
        }
        
        # If not confirmed, return preview
        if not confirmation:
            return {
                "action": "preview",
                "message": "This is a preview. Set confirmation=true to actually delete.",
                **summary,
                "warning": "⚠️ This action cannot be undone!"
            }
        
        # CONFIRMED - Actually delete
        logger.warning(f"🗑️ DELETING CLIENT: {client_data.get('company_name')} ({client_id})")
        
        # Delete associated data (cascade should handle most, but be explicit)
        supabase.table("opportunities").delete().eq("client_id", client_id).execute()
        supabase.table("document_uploads").delete().eq("client_id", client_id).execute()
        supabase.table("content_calendars").delete().eq("client_id", client_id).execute()
        supabase.table("client_subreddit_config").delete().eq("client_id", client_id).execute()
        supabase.table("client_keyword_config").delete().eq("client_id", client_id).execute()
        
        # Delete client
        supabase.table("clients").delete().eq("client_id", client_id).execute()
        
        logger.info(f"✅ Client deleted: {client_data.get('company_name')}")
        
        return {
            "success": True,
            "action": "deleted",
            "message": f"Client '{client_data.get('company_name')}' and all associated data deleted",
            **summary
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting client: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/clients/bulk-delete")
async def bulk_delete_clients(client_ids: List[str], confirmation: bool = False):
    """
    Delete multiple clients at once
    
    Requires confirmation=true to actually delete
    """
    try:
        results = []
        
        for client_id in client_ids:
            try:
                result = await delete_client(client_id, confirmation)
                results.append({
                    "client_id": client_id,
                    "status": "success",
                    "result": result
                })
            except Exception as e:
                results.append({
                    "client_id": client_id,
                    "status": "failed",
                    "error": str(e)
                })
        
        return {
            "success": True,
            "deleted": len([r for r in results if r["status"] == "success"]),
            "failed": len([r for r in results if r["status"] == "failed"]),
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Bulk delete error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/resend-notification/{client_id}")
async def resend_notification(client_id: str):
    """
    Manually resend onboarding notification email
    Useful for testing or if initial email failed
    """
    try:
        import os
        from services.onboarding_orchestrator import OnboardingOrchestrator
        
        supabase = get_supabase()
        openai_key = os.getenv("OPENAI_API_KEY")
        
        # Get client
        client = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        if not client.data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        # Send notification
        orchestrator = OnboardingOrchestrator(supabase, openai_key)
        result = await orchestrator._send_welcome_email(client.data[0], {"success": True, "items": 0})
        
        return {
            "success": result.get("success"),
            "client_id": client_id,
            "email": client.data[0].get("notification_email"),
            "result": result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error resending notification: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/send-weekly-reports")
async def send_weekly_reports_to_all():
    """
    Manually trigger weekly report generation for all clients
    Useful for testing the report system
    """
    try:
        import asyncio
        from workers.weekly_report_generator import WeeklyReportGenerator
        
        logger.info("Manual weekly report generation triggered")
        
        generator = WeeklyReportGenerator()
        result = await generator.send_reports_to_all_clients()
        
        return {
            "success": True,
            "result": result
        }
        
    except Exception as e:
        logger.error(f"Error sending weekly reports: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/trigger-orchestrator/{client_id}")
async def trigger_full_orchestrator(client_id: str, background_tasks: BackgroundTasks):
    """
    Manually trigger full orchestrator for a client
    
    This will:
    1. Run AUTO_IDENTIFY keyword expansion
    2. Scrape Reddit for 50-100 opportunities
    3. Score and prioritize opportunities
    4. Generate Intelligence Report & Sample Content
    5. Send welcome email with reports
    
    Use this when:
    - Initial onboarding failed to collect opportunities
    - Client has 0 opportunities in database
    - Need to re-run full data collection
    
    Returns immediately, processing runs in background (5-10 minutes)
    """
    try:
        from services.onboarding_orchestrator import OnboardingOrchestrator
        from services.delayed_report_workflow import DelayedReportWorkflow
        
        supabase = get_supabase()
        openai_key = os.getenv("OPENAI_API_KEY")
        
        # Fetch client
        client_response = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        
        if not client_response.data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        client = client_response.data[0]
        company_name = client.get('company_name', 'Client')
        
        logger.info(f"🚀 Manual orchestrator trigger for: {company_name}")
        
        # Run orchestrator in background
        async def run_full_workflow():
            try:
                # Step 1: Run orchestrator (scraping + scoring)
                orchestrator = OnboardingOrchestrator(supabase, openai_key)
                result = await orchestrator.process_client_onboarding(client_id)
                
                logger.info(f"✅ Orchestrator completed for {company_name}: {result}")
                
                # Step 2: Run delayed report workflow
                openai_client = AsyncOpenAI(api_key=openai_key)
                workflow = DelayedReportWorkflow(supabase, openai_client)
                
                notification_email = client.get('primary_contact_email') or client.get('notification_email')
                slack_webhook = client.get('slack_webhook_url')
                
                await workflow.run_workflow(
                    client_id=client_id,
                    notification_email=notification_email,
                    slack_webhook=slack_webhook,
                    min_opportunities=10,
                    timeout_seconds=600
                )
                
                logger.info(f"✅ Full workflow completed for {company_name}")
                
            except Exception as e:
                logger.error(f"❌ Workflow error for {company_name}: {str(e)}", exc_info=True)
        
        # Add to background tasks
        background_tasks.add_task(run_full_workflow)
        
        return {
            "success": True,
            "message": f"Full orchestrator triggered for {company_name}",
            "client_id": client_id,
            "company_name": company_name,
            "estimated_completion": "5-10 minutes",
            "steps": [
                "1. AUTO_IDENTIFY keywords",
                "2. Scrape Reddit for opportunities",
                "3. Score and prioritize",
                "4. Generate Intelligence Report",
                "5. Generate Sample Content",
                "6. Send welcome email"
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error triggering orchestrator: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/send-weekly-report/{client_id}")
async def send_weekly_report_to_client(client_id: str):
    """
    Send weekly report to a specific client
    Useful for testing with individual clients
    """
    try:
        from workers.weekly_report_generator import WeeklyReportGenerator
        
        supabase = get_supabase()
        
        # Fetch client
        client_response = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        
        if not client_response.data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        client = client_response.data[0]
        
        logger.info(f"Manual weekly report for: {client.get('company_name')}")
        
        generator = WeeklyReportGenerator()
        result = await generator._generate_and_send_report(client)
        
        return {
            "success": result.get("success"),
            "client_id": client_id,
            "company_name": client.get("company_name"),
            "result": result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error sending weekly report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/regenerate-reports/{client_id}")
async def regenerate_onboarding_reports(client_id: str, background_tasks: BackgroundTasks):
    """
    Regenerate Intelligence Report and Sample Content for a client
    
    Use this when:
    - Initial onboarding workflow failed
    - Client needs updated reports
    - Testing report generation
    
    Returns immediately, reports generated in background
    """
    try:
        from fastapi import BackgroundTasks
        supabase = get_supabase()
        
        # Verify client exists
        client_response = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        
        if not client_response.data:
            raise HTTPException(status_code=404, detail=f"Client {client_id} not found")
        
        client = client_response.data[0]
        company_name = client.get("company_name")
        notification_email = client.get("primary_contact_email") or client.get("notification_email")
        slack_webhook = client.get("slack_webhook_url")
        
        if not notification_email:
            raise HTTPException(status_code=400, detail="Client has no email address for reports")
        
        logger.info(f"🔄 Regenerating reports for: {company_name} ({client_id})")
        
        # Trigger delayed report workflow in background
        from client_onboarding_router import run_onboarding_with_delayed_reports
        
        background_tasks.add_task(
            run_onboarding_with_delayed_reports,
            client_id,
            notification_email,
            slack_webhook
        )
        
        return {
            "success": True,
            "message": f"Report generation started for {company_name}",
            "client_id": client_id,
            "email": notification_email,
            "estimated_time": "5-10 minutes",
            "note": "You will receive Intelligence Report and Sample Content via email"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error regenerating reports: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/trigger-reddit-scan")
async def trigger_reddit_scan(background_tasks: BackgroundTasks, request: dict = None):
    """
    Manually trigger Reddit opportunity scan for all clients or specific client
    
    Used for:
    - Immediate scanning after onboarding (don't wait for scheduled job)
    - Testing Reddit monitoring
    - Recovering from failed scans
    
    Body (optional):
    {
        "client_id": "uuid"  # If omitted, scans all active clients
    }
    """
    try:
        client_id = request.get("client_id") if request else None
        
        if client_id:
            # Scan specific client
            supabase = get_supabase()
            client = supabase.table("clients").select("*").eq("client_id", client_id).execute()
            
            if not client.data:
                raise HTTPException(status_code=404, detail="Client not found")
            
            client_data = client.data[0]
            company_name = client_data.get("company_name")
            subreddits = client_data.get("target_subreddits", [])
            keywords = client_data.get("target_keywords", [])
            
            if not subreddits or not keywords:
                raise HTTPException(
                    status_code=400,
                    detail=f"Client {company_name} has no subreddits or keywords configured"
                )
            
            logger.info(f"🔍 Triggering Reddit scan for: {company_name}")
            logger.info(f"   Subreddits: {subreddits}")
            logger.info(f"   Keywords: {keywords}")
            
            # Run scan in background
            from workers.brand_mention_monitor import scan_for_opportunities, save_opportunities
            
            def scan_and_save():
                opportunities = scan_for_opportunities(client_id, company_name, subreddits, keywords)
                if opportunities:
                    save_opportunities(opportunities)
                    logger.info(f"✅ Created {len(opportunities)} opportunities for {company_name}")
                else:
                    logger.warning(f"⚠️  No opportunities found for {company_name}")
                return opportunities
            
            background_tasks.add_task(scan_and_save)
            
            return {
                "success": True,
                "message": f"Reddit scan started for {company_name}",
                "client_id": client_id,
                "estimated_time": "2-3 minutes",
                "note": "Check dashboard for opportunities"
            }
        
        else:
            # Scan all active clients
            logger.info("🔍 Triggering Reddit scan for ALL active clients")
            
            from workers.brand_mention_monitor import run_opportunity_monitor
            
            background_tasks.add_task(run_opportunity_monitor)
            
            return {
                "success": True,
                "message": "Reddit scan started for all active clients",
                "estimated_time": "5-10 minutes",
                "note": "Scanning all configured subreddits"
            }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error triggering Reddit scan: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/run-worker-pipeline")
async def run_worker_pipeline(background_tasks: BackgroundTasks, request: dict):
    """
    Manually trigger the full worker pipeline for a client
    
    Pipeline stages:
    1. Opportunity Scoring
    2. Product Matchback
    3. Content Generation
    4. Voice Application
    
    Used for:
    - Processing newly discovered opportunities
    - Regenerating content after document updates
    - Testing the full pipeline
    
    Body:
    {
        "client_id": "uuid",
        "force_regenerate": false  # Optional: regenerate existing content
    }
    """
    try:
        client_id = request.get("client_id")
        force_regenerate = request.get("force_regenerate", False)
        
        if not client_id:
            raise HTTPException(status_code=400, detail="client_id is required")
        
        supabase = get_supabase()
        client = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        
        if not client.data:
            raise HTTPException(status_code=404, detail="Client not found")
        
        company_name = client.data[0].get("company_name")
        
        logger.info(f"⚙️  Triggering worker pipeline for: {company_name}")
        
        # Run pipeline in background
        from workers.scheduler import run_full_pipeline
        
        background_tasks.add_task(
            run_full_pipeline,
            client_id,
            force_regenerate
        )
        
        return {
            "success": True,
            "message": f"Worker pipeline started for {company_name}",
            "client_id": client_id,
            "force_regenerate": force_regenerate,
            "estimated_time": "5-10 minutes",
            "pipeline_stages": [
                "Opportunity Scoring",
                "Product Matchback",
                "Content Generation",
                "Voice Application"
            ],
            "note": "Check dashboard for content pieces after completion"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error running worker pipeline: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class TestContentRequest(BaseModel):
    client_id: str = None
    count: int = 2
    opportunity_id: str = None  # Optional: Test specific opportunity


@router.post("/test-content-generation")
async def test_content_generation_sync(request: TestContentRequest = None):
    """
    Synchronous test endpoint for content generation.
    Runs content generation directly (not in background) so we can see errors.

    This is for debugging only - returns full result including any errors.

    Body (optional):
    {
        "client_id": "uuid",  # If omitted, uses first active client
        "count": 2            # Number of opportunities to process (default 2)
    }
    """
    import traceback

    try:
        logger.info("🧪 TEST CONTENT GENERATION - Starting synchronous test")

        # Import the worker
        from workers.content_generation_worker import ContentGenerationWorker

        # Create worker instance
        worker = ContentGenerationWorker()

        # Get client_id from request or find first active client
        client_id = request.client_id if request and request.client_id else None
        count = request.count if request and request.count else 2

        if not client_id:
            # Find first active client with opportunities
            clients = worker.supabase.table("clients").select("client_id, company_name").eq("subscription_status", "active").limit(5).execute()
            for c in (clients.data or []):
                opps = worker.supabase.table("opportunities").select("opportunity_id").eq("client_id", c["client_id"]).limit(1).execute()
                if opps.data:
                    client_id = c["client_id"]
                    logger.info(f"🧪 Using client: {c['company_name']} ({client_id})")
                    break

            if not client_id:
                return {"success": False, "error": "No active clients with opportunities found"}

        logger.info(f"🧪 Calling process_all_opportunities for client {client_id}")

        # Check if specific opportunity_id was provided
        specific_opp_id = request.opportunity_id if request and request.opportunity_id else None

        if specific_opp_id:
            # Get specific opportunity
            logger.info(f"🧪 Testing specific opportunity: {specific_opp_id}")
            worker_query = worker.supabase.table("opportunities")\
                .select("opportunity_id, client_id, thread_title, original_post_text, subreddit, thread_url, date_found")\
                .eq("opportunity_id", specific_opp_id)\
                .execute()
            test_opps = worker_query.data if worker_query.data else []
            opps_count = len(test_opps)
            sample_opp = test_opps[0] if test_opps else None
            worker_query_count = len(test_opps)
        else:
            # First, let's check how many opportunities exist using worker's supabase
            opps_check = worker.supabase.table("opportunities").select("opportunity_id, thread_title").eq("client_id", client_id).order("date_found", desc=True).limit(5).execute()
            opps_count = len(opps_check.data) if opps_check.data else 0
            sample_opp = opps_check.data[0] if opps_check.data else None

            # Also check what the worker's query returns
            worker_query = worker.supabase.table("opportunities").select("opportunity_id, client_id, thread_title, original_post_text, subreddit, thread_url, date_found").eq("client_id", client_id).order("date_found", desc=True).limit(10).execute()
            worker_query_count = len(worker_query.data) if worker_query.data else 0

            # Directly call generate_content_for_client with limited opportunities to avoid timeout
            # Take first 'count' opportunities for quick test
            test_opps = worker_query.data[:count] if worker_query.data else []

        if test_opps:
            try:
                result = worker.generate_content_for_client(
                    client_id=client_id,
                    opportunities=test_opps,
                    delivery_batch=f"TEST-{datetime.now().strftime('%Y-%m-%d')}"
                )
            except Exception as gen_error:
                gen_tb = traceback.format_exc()
                result = {
                    "success": False,
                    "error": f"Generation failed: {str(gen_error)}",
                    "traceback": gen_tb
                }
        else:
            result = {"success": False, "error": "No opportunities found"}

        logger.info(f"🧪 Result: {result}")

        # Check actual content in database after generation
        content_check = worker.supabase.table("content_delivered").select("id, client_id, content_type, subreddit_name, delivered_at").eq("client_id", client_id).order("delivered_at", desc=True).limit(5).execute()
        content_count = len(content_check.data) if content_check.data else 0
        sample_content = content_check.data[0] if content_check.data else None

        return {
            "success": True,
            "test_type": "synchronous_content_generation",
            "client_id": client_id,
            "opportunities_check": {
                "count": opps_count,
                "sample": sample_opp,
                "worker_query_count": worker_query_count,
                "test_opportunities": [
                    {
                        "id": o.get("opportunity_id"),
                        "subreddit": o.get("subreddit"),
                        "title": o.get("thread_title", "")[:80]
                    } for o in test_opps
                ]
            },
            "result": result,
            "content_in_database": {
                "count": content_count,
                "sample": sample_content
            }
        }

    except Exception as e:
        error_traceback = traceback.format_exc()
        logger.error(f"🧪 TEST FAILED: {e}")
        logger.error(f"Traceback: {error_traceback}")

        return {
            "success": False,
            "test_type": "synchronous_content_generation",
            "error": str(e),
            "traceback": error_traceback
        }


@router.post("/run-opportunity-scoring")
async def run_opportunity_scoring(client_id: str = None):
    """
    Manually trigger opportunity scoring for all or specific client.
    Scores opportunities by: buying intent, pain points, questions, engagement, urgency.
    """
    try:
        logger.info(f"🎯 Starting opportunity scoring (client_id: {client_id or 'all'})")

        from workers.opportunity_scoring_worker import OpportunityScoringWorker

        worker = OpportunityScoringWorker()
        result = worker.process_all_opportunities(client_id)

        logger.info(f"🎯 Scoring complete: {result}")

        return {
            "success": True,
            "action": "opportunity_scoring",
            "client_id": client_id,
            "result": result
        }

    except Exception as e:
        logger.error(f"🎯 Scoring failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }


@router.get("/pipeline-status")
async def get_pipeline_status():
    """
    Get status of all pipeline components.
    """
    try:
        supabase = get_supabase()

        # Check voice profiles
        voice_profiles = supabase.table("voice_profiles").select("id", count="exact").execute()
        voice_count = voice_profiles.count if voice_profiles.count else 0

        # Check scored opportunities
        scored_opps = supabase.table("opportunities").select("opportunity_id", count="exact").not_.is_("opportunity_score", "null").execute()
        scored_count = scored_opps.count if scored_opps.count else 0

        # Check total opportunities
        total_opps = supabase.table("opportunities").select("opportunity_id", count="exact").execute()
        total_count = total_opps.count if total_opps.count else 0

        # Check knowledge base (document_uploads or client_documents)
        try:
            docs = supabase.table("document_uploads").select("id", count="exact").execute()
            docs_count = docs.count if docs.count else 0
        except:
            docs_count = 0

        # Check vector embeddings
        try:
            embeddings = supabase.table("vector_embeddings").select("id", count="exact").execute()
            embeddings_count = embeddings.count if embeddings.count else 0
        except:
            embeddings_count = 0

        # Check content delivered
        content = supabase.table("content_delivered").select("id", count="exact").execute()
        content_count = content.count if content.count else 0

        return {
            "pipeline_components": {
                "voice_database": {
                    "profiles": voice_count,
                    "status": "POPULATED" if voice_count > 0 else "EMPTY"
                },
                "opportunity_scoring": {
                    "scored": scored_count,
                    "total": total_count,
                    "percentage": round(scored_count / total_count * 100, 1) if total_count > 0 else 0,
                    "status": "COMPLETE" if scored_count == total_count and total_count > 0 else "PARTIAL" if scored_count > 0 else "NOT_RUN"
                },
                "knowledge_base": {
                    "documents": docs_count,
                    "embeddings": embeddings_count,
                    "status": "POPULATED" if embeddings_count > 0 else "EMPTY"
                },
                "content_generation": {
                    "pieces": content_count,
                    "status": "ACTIVE" if content_count > 0 else "NO_CONTENT"
                }
            },
            "overall_status": "READY" if voice_count > 0 and scored_count > 0 else "NEEDS_SETUP"
        }

    except Exception as e:
        logger.error(f"Pipeline status check failed: {e}")
        return {"error": str(e)}


@router.post("/configure-subreddits")
async def configure_client_subreddits(request: dict):
    """
    Add subreddits to client_subreddit_config table for an existing client.
    Use this to fix clients that were onboarded without proper subreddit configuration.

    Request body:
    {
        "client_id": "uuid-here",
        "subreddits": ["Menopause", "PCOS", "TryingForABaby"]
    }
    """
    try:
        supabase = get_supabase()
        client_id = request.get("client_id")
        subreddits = request.get("subreddits", [])

        if not client_id:
            return {"success": False, "error": "client_id is required"}
        if not subreddits:
            return {"success": False, "error": "subreddits list is required"}

        # Verify client exists
        client = supabase.table("clients").select("client_id, company_name").eq("client_id", client_id).execute()
        if not client.data:
            return {"success": False, "error": f"Client {client_id} not found"}

        # Check existing subreddit configs
        existing = supabase.table("client_subreddit_config").select("subreddit_name").eq("client_id", client_id).execute()
        existing_names = [s["subreddit_name"].lower() for s in (existing.data or [])]

        # Insert new subreddits (skip duplicates)
        new_configs = []
        skipped = []
        for sub in subreddits:
            sub_clean = sub.lower().replace("r/", "")
            if sub_clean in existing_names:
                skipped.append(sub_clean)
            else:
                new_configs.append({
                    "client_id": client_id,
                    "subreddit_name": sub_clean,
                    "is_active": True,
                    "created_at": datetime.utcnow().isoformat()
                })

        if new_configs:
            supabase.table("client_subreddit_config").insert(new_configs).execute()
            logger.info(f"Added {len(new_configs)} subreddits for client {client_id}")

        return {
            "success": True,
            "client_id": client_id,
            "client_name": client.data[0].get("company_name"),
            "added": [c["subreddit_name"] for c in new_configs],
            "skipped_duplicates": skipped,
            "total_configured": len(existing_names) + len(new_configs)
        }

    except Exception as e:
        logger.error(f"Configure subreddits failed: {e}")
        return {"success": False, "error": str(e)}


@router.get("/client-subreddits/{client_id}")
async def get_client_subreddits(client_id: str):
    """Get all configured subreddits for a client"""
    try:
        supabase = get_supabase()

        # Get subreddit configs
        configs = supabase.table("client_subreddit_config").select("*").eq("client_id", client_id).execute()

        return {
            "client_id": client_id,
            "subreddits": configs.data or [],
            "count": len(configs.data or [])
        }

    except Exception as e:
        logger.error(f"Get subreddits failed: {e}")
        return {"error": str(e)}


@router.post("/reprocess-documents/{client_id}")
async def reprocess_stuck_documents(client_id: str, background_tasks: BackgroundTasks):
    """
    Reprocess documents stuck in 'processing' status

    Use this when documents were uploaded but chunking/embedding failed
    (typically due to timeout)

    Runs in background - returns immediately
    """
    try:
        supabase = get_supabase()

        # Get client info
        client = supabase.table("clients").select("company_name").eq("client_id", client_id).execute()
        if not client.data:
            raise HTTPException(status_code=404, detail="Client not found")

        company_name = client.data[0].get("company_name")

        # Get stuck documents
        stuck_docs = supabase.table("document_uploads")\
            .select("*")\
            .eq("client_id", client_id)\
            .eq("processing_status", "processing")\
            .execute()

        if not stuck_docs.data:
            return {
                "success": True,
                "message": "No stuck documents to reprocess",
                "client_id": client_id
            }

        logger.info(f"🔄 Reprocessing {len(stuck_docs.data)} stuck documents for {company_name}")

        async def reprocess_documents():
            """Background task to reprocess documents"""
            from services.document_ingestion_service import DocumentIngestionService

            service = DocumentIngestionService(supabase, os.getenv("OPENAI_API_KEY"))
            results = []

            for doc in stuck_docs.data:
                doc_id = doc.get("id")
                filename = doc.get("filename")

                try:
                    # Mark as failed first (so we can retry)
                    supabase.table("document_uploads").update({
                        "processing_status": "retrying"
                    }).eq("id", doc_id).execute()

                    # We don't have the original file content, so mark as failed
                    # Future: store files in Supabase Storage for retry
                    supabase.table("document_uploads").update({
                        "processing_status": "failed",
                        "error_message": "Cannot reprocess - original file not stored. Please re-upload."
                    }).eq("id", doc_id).execute()

                    results.append({
                        "document_id": doc_id,
                        "filename": filename,
                        "status": "marked_failed",
                        "message": "Please re-upload this file"
                    })

                except Exception as e:
                    logger.error(f"Error marking document {doc_id}: {e}")
                    results.append({
                        "document_id": doc_id,
                        "filename": filename,
                        "status": "error",
                        "error": str(e)
                    })

            logger.info(f"✅ Reprocessing complete for {company_name}: {len(results)} documents updated")
            return results

        background_tasks.add_task(reprocess_documents)

        return {
            "success": True,
            "message": f"Reprocessing {len(stuck_docs.data)} stuck documents for {company_name}",
            "client_id": client_id,
            "documents": [{"id": d["id"], "filename": d.get("filename")} for d in stuck_docs.data],
            "note": "Documents will be marked as failed - please re-upload them"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Reprocess documents failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/documents/{client_id}/stuck")
async def delete_stuck_documents(client_id: str):
    """
    Delete all documents stuck in 'processing' status

    Use this to clean up failed uploads so user can try again
    """
    try:
        supabase = get_supabase()

        # Get client info
        client = supabase.table("clients").select("company_name").eq("client_id", client_id).execute()
        if not client.data:
            raise HTTPException(status_code=404, detail="Client not found")

        company_name = client.data[0].get("company_name")

        # Get stuck documents before deleting
        stuck_docs = supabase.table("document_uploads")\
            .select("id, filename")\
            .eq("client_id", client_id)\
            .eq("processing_status", "processing")\
            .execute()

        if not stuck_docs.data:
            return {
                "success": True,
                "message": "No stuck documents to delete",
                "client_id": client_id
            }

        # Delete stuck documents
        deleted_count = len(stuck_docs.data)
        supabase.table("document_uploads")\
            .delete()\
            .eq("client_id", client_id)\
            .eq("processing_status", "processing")\
            .execute()

        logger.info(f"🗑️ Deleted {deleted_count} stuck documents for {company_name}")

        return {
            "success": True,
            "message": f"Deleted {deleted_count} stuck documents for {company_name}",
            "client_id": client_id,
            "deleted": [{"id": d["id"], "filename": d.get("filename")} for d in stuck_docs.data]
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete stuck documents failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rag-diagnostic/{client_id}")
async def diagnose_rag_system(client_id: str):
    """
    Diagnostic endpoint to check if RAG (knowledge base) is working for a client.

    Checks:
    1. Document embeddings table has data
    2. RPC function match_knowledge_embeddings exists
    3. Test query returns results
    """
    try:
        supabase = get_supabase()
        results = {
            "client_id": client_id,
            "checks": {}
        }

        # Check 1: Does document_embeddings table have data for this client?
        try:
            embeddings = supabase.table("document_embeddings")\
                .select("id, document_id, chunk_index, created_at")\
                .eq("client_id", client_id)\
                .limit(5)\
                .execute()
            results["checks"]["document_embeddings"] = {
                "status": "found" if embeddings.data else "empty",
                "count": len(embeddings.data) if embeddings.data else 0,
                "sample": embeddings.data[:2] if embeddings.data else []
            }
        except Exception as e:
            results["checks"]["document_embeddings"] = {
                "status": "error",
                "error": str(e)
            }

        # Check 2: Does document_uploads table have completed docs?
        try:
            docs = supabase.table("document_uploads")\
                .select("id, filename, processing_status, chunk_count")\
                .eq("client_id", client_id)\
                .execute()
            results["checks"]["document_uploads"] = {
                "status": "found" if docs.data else "empty",
                "count": len(docs.data) if docs.data else 0,
                "completed": len([d for d in (docs.data or []) if d.get("processing_status") == "completed"]),
                "total_chunks_reported": sum(d.get("chunk_count", 0) or 0 for d in (docs.data or []))
            }
        except Exception as e:
            results["checks"]["document_uploads"] = {
                "status": "error",
                "error": str(e)
            }

        # Check 3: Does the RPC function exist?
        try:
            test_embedding = [0.0] * 1536
            rpc_result = supabase.rpc(
                'match_knowledge_embeddings',
                {
                    'query_embedding': test_embedding,
                    'client_id': client_id,
                    'similarity_threshold': 0.0,  # Very low to get any results
                    'match_count': 3
                }
            ).execute()
            results["checks"]["rpc_function"] = {
                "status": "exists",
                "test_results_count": len(rpc_result.data) if rpc_result.data else 0
            }
        except Exception as e:
            error_msg = str(e).lower()
            if "does not exist" in error_msg or "function" in error_msg:
                results["checks"]["rpc_function"] = {
                    "status": "missing",
                    "error": "Function match_knowledge_embeddings not found",
                    "fix": "Run sql/match_knowledge_embeddings.sql in Supabase SQL Editor"
                }
            else:
                results["checks"]["rpc_function"] = {
                    "status": "error",
                    "error": str(e)
                }

        # Summary
        embeddings_ok = results["checks"].get("document_embeddings", {}).get("status") == "found"
        rpc_ok = results["checks"].get("rpc_function", {}).get("status") == "exists"

        if embeddings_ok and rpc_ok:
            results["rag_status"] = "WORKING"
            results["message"] = "RAG system is configured and has embeddings"
        elif rpc_ok and not embeddings_ok:
            results["rag_status"] = "NO_EMBEDDINGS"
            results["message"] = "RPC function exists but no embeddings found - documents may not have been vectorized"
        elif not rpc_ok:
            results["rag_status"] = "RPC_MISSING"
            results["message"] = "match_knowledge_embeddings function not found - run migration SQL"
        else:
            results["rag_status"] = "NOT_CONFIGURED"
            results["message"] = "RAG system needs setup"

        return results

    except Exception as e:
        logger.error(f"RAG diagnostic failed: {e}")
        return {"error": str(e)}


@router.post("/test-voice-crawl")
async def test_voice_crawl_single_subreddit(request: dict, background_tasks: BackgroundTasks):
    """
    Test voice crawl for a SINGLE subreddit (runs in background).

    Body:
    {
        "client_id": "uuid",
        "subreddit": "HomeImprovement"
    }

    This crawls just one subreddit with reduced limits for testing.
    Returns immediately, crawl runs in background.
    """
    try:
        client_id = request.get("client_id")
        subreddit = request.get("subreddit")

        if not client_id or not subreddit:
            return {"error": "client_id and subreddit are required"}

        # Clean subreddit name
        subreddit = subreddit.replace("r/", "").strip()

        logger.info(f"🎤 Starting background voice crawl for r/{subreddit} (client: {client_id})")

        async def run_voice_crawl():
            try:
                from workers.voice_database_worker import VoiceDatabaseWorker

                worker = VoiceDatabaseWorker()
                worker.TOP_USERS_PER_SUBREDDIT = 10  # Very small for quick test
                worker.COMMENTS_PER_USER = 5

                await worker.analyze_subreddit_voice(subreddit, client_id)
                logger.info(f"✅ Voice crawl complete for r/{subreddit}")
            except Exception as e:
                logger.error(f"Background voice crawl failed: {e}")

        background_tasks.add_task(run_voice_crawl)

        return {
            "success": True,
            "message": f"Voice crawl started in background for r/{subreddit}",
            "client_id": client_id,
            "subreddit": subreddit,
            "note": "Check /api/admin/pipeline-status in ~60 seconds to see if profile was created"
        }

    except Exception as e:
        logger.error(f"Voice crawl test failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }


@router.post("/create-voice-profile")
async def create_voice_profile_manual(request: dict):
    """
    Create a voice profile manually (without Reddit crawling).

    This is useful for testing the pipeline when Reddit crawling is slow.

    Body:
    {
        "client_id": "uuid",
        "subreddit": "HomeImprovement"
    }

    v3: Uses delete+insert instead of upsert
    """
    try:
        supabase = get_supabase()

        client_id = request.get("client_id")
        subreddit = request.get("subreddit")

        if not client_id or not subreddit:
            return {"error": "client_id and subreddit are required"}

        subreddit = subreddit.replace("r/", "").strip().lower()

        # Default voice profile values
        profile_values = {
            "dominant_tone": "helpful",
            "formality_score": 0.3,
            "lowercase_start_pct": 25,
            "exclamation_usage_pct": 12
        }

        # Use a synthetic redditor_username for subreddit-wide voice (table requires it)
        synthetic_username = f"__subreddit_voice_{subreddit}__"

        # Try multiple column combinations until one works
        column_attempts = [
            # Attempt 1: Most complete
            {
                "client_id": client_id,
                "redditor_username": synthetic_username,
                "subreddit": subreddit,
                "dominant_tone": profile_values["dominant_tone"],
                "formality_score": profile_values["formality_score"],
                "lowercase_start_pct": profile_values["lowercase_start_pct"],
                "exclamation_usage_pct": profile_values["exclamation_usage_pct"],
                "created_at": datetime.utcnow().isoformat()
            },
            # Attempt 2: Without numeric columns
            {
                "client_id": client_id,
                "redditor_username": synthetic_username,
                "subreddit": subreddit,
                "dominant_tone": profile_values["dominant_tone"],
                "created_at": datetime.utcnow().isoformat()
            },
            # Attempt 3: Just the required columns
            {
                "client_id": client_id,
                "redditor_username": synthetic_username,
                "subreddit": subreddit,
                "created_at": datetime.utcnow().isoformat()
            }
        ]

        inserted_data = None
        last_error = None

        # First, try to delete any existing profile for this client/subreddit
        try:
            supabase.table("voice_profiles").delete().eq("client_id", client_id).eq("subreddit", subreddit).execute()
            logger.info(f"Deleted existing voice profile for r/{subreddit}")
        except Exception as e:
            logger.warning(f"No existing profile to delete or delete failed: {e}")

        for i, data in enumerate(column_attempts):
            try:
                # Use insert instead of upsert (since we just deleted any existing)
                supabase.table("voice_profiles").insert(data).execute()
                inserted_data = data
                logger.info(f"✅ Created voice profile with attempt {i+1} for r/{subreddit}")
                break
            except Exception as e:
                last_error = str(e)
                logger.warning(f"Voice profile attempt {i+1} failed: {e}")
                continue

        if not inserted_data:
            return {
                "success": False,
                "error": f"All column combinations failed. Last error: {last_error}"
            }

        return {
            "success": True,
            "version": "v3",
            "client_id": client_id,
            "subreddit": subreddit,
            "profile_created": True,
            "columns_used": list(inserted_data.keys()),
            "profile_values": profile_values
        }

    except Exception as e:
        logger.error(f"Manual voice profile creation failed: {e}")
        return {
            "version": "v3",
            "success": False,
            "error": str(e)
        }


@router.post("/test-rag-search")
async def test_rag_search(request: dict):
    """
    Test RAG knowledge base search with a query.

    Body:
    {
        "client_id": "uuid",
        "query": "How do electric fireplaces compare to gas?",
        "threshold": 0.50
    }
    """
    try:
        client_id = request.get("client_id")
        query = request.get("query", "electric fireplace heating efficiency")
        threshold = request.get("threshold", 0.50)

        if not client_id:
            return {"error": "client_id required"}

        # Direct test with embedding
        from openai import OpenAI
        import os

        openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        supabase = get_supabase()

        # Generate embedding - MUST match model used for stored embeddings
        embedding_response = openai_client.embeddings.create(
            model="text-embedding-ada-002",  # Same as populate_embeddings.py
            input=query[:8000]
        )
        query_embedding = embedding_response.data[0].embedding
        embedding_dim = len(query_embedding)

        # Direct RPC call
        rpc_result = supabase.rpc(
            'match_knowledge_embeddings',
            {
                'query_embedding': query_embedding,
                'client_id': client_id,
                'similarity_threshold': threshold,
                'match_count': 5
            }
        ).execute()

        # Also test zero threshold to verify function works
        rpc_result_zero = supabase.rpc(
            'match_knowledge_embeddings',
            {
                'query_embedding': query_embedding,
                'client_id': client_id,
                'similarity_threshold': 0.0,  # Get ALL matches
                'match_count': 5
            }
        ).execute()

        return {
            "success": True,
            "client_id": client_id,
            "query": query,
            "threshold": threshold,
            "embedding_dimensions": embedding_dim,
            "insights_found_at_threshold": len(rpc_result.data) if rpc_result.data else 0,
            "insights_found_at_zero": len(rpc_result_zero.data) if rpc_result_zero.data else 0,
            "insights_at_threshold": rpc_result.data or [],
            "insights_at_zero": [
                {
                    "chunk_text": r.get("chunk_text", "")[:200],
                    "similarity": r.get("similarity")
                } for r in (rpc_result_zero.data or [])
            ]
        }

    except Exception as e:
        logger.error(f"RAG search test failed: {e}")
        import traceback
        return {"success": False, "error": str(e), "traceback": traceback.format_exc()}


@router.get("/search-opportunities/{client_id}")
async def search_opportunities(client_id: str, keywords: str = None, limit: int = 20):
    """
    Search opportunities for a client by keywords.

    Use this to find product-relevant opportunities for testing.

    Args:
        client_id: Client UUID
        keywords: Comma-separated keywords to search (e.g., "fireplace,tv lift,heating")
        limit: Max results (default 20)

    Example:
        GET /api/admin/search-opportunities/999ac53f-...?keywords=fireplace,electric,tv&limit=10
    """
    try:
        supabase = get_supabase()

        # Get all opportunities for client
        query = supabase.table("opportunities")\
            .select("opportunity_id, thread_title, subreddit, opportunity_score, original_post_text, date_found, matched_keywords, thread_url, status")\
            .eq("client_id", client_id)\
            .order("date_found", desc=True)\
            .limit(500)  # Get more to filter

        result = query.execute()

        if not result.data:
            return {"success": True, "count": 0, "opportunities": [], "message": "No opportunities found"}

        opportunities = result.data

        # Filter by keywords if provided
        if keywords:
            keyword_list = [k.strip().lower() for k in keywords.split(",")]
            filtered = []
            for opp in opportunities:
                title = (opp.get("thread_title") or "").lower()
                content = (opp.get("original_post_text") or "").lower()
                combined = f"{title} {content}"

                for kw in keyword_list:
                    if kw in combined:
                        opp["matched_keyword"] = kw
                        filtered.append(opp)
                        break

            opportunities = filtered

        # Limit results
        opportunities = opportunities[:limit]

        return {
            "success": True,
            "client_id": client_id,
            "keywords_searched": keywords,
            "count": len(opportunities),
            "opportunities": [
                {
                    "opportunity_id": o.get("opportunity_id"),
                    "thread_title": o.get("thread_title"),
                    "subreddit": o.get("subreddit"),
                    "score": o.get("opportunity_score"),
                    "matched_keyword": o.get("matched_keyword"),  # From keyword filter
                    "matched_keywords": o.get("matched_keywords"),  # From DB (JSON string)
                    "thread_url": o.get("thread_url"),
                    "status": o.get("status"),
                    "date_found": o.get("date_found")
                } for o in opportunities
            ]
        }

    except Exception as e:
        logger.error(f"Search opportunities failed: {e}")
        return {"success": False, "error": str(e)}


@router.post("/reprocess-embeddings/{client_id}")
async def reprocess_document_embeddings(client_id: str):
    """
    Reprocess documents from document_uploads and generate embeddings in document_embeddings.

    This fixes the gap where documents were uploaded but embeddings weren't created.

    Process:
    1. Get all completed documents from document_uploads
    2. Try to get chunks from document_chunks OR vector_embeddings
    3. Generate OpenAI embeddings for each chunk
    4. Insert into document_embeddings table (which RAG queries)
    """
    from openai import OpenAI
    import os

    try:
        supabase = get_supabase()
        openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        # Get client info
        client = supabase.table("clients").select("company_name").eq("client_id", client_id).execute()
        if not client.data:
            return {"error": f"Client {client_id} not found"}
        company_name = client.data[0].get("company_name")

        logger.info(f"🔄 Reprocessing embeddings for {company_name} ({client_id})")

        # Get completed documents from document_uploads
        docs = supabase.table("document_uploads")\
            .select("id, filename, file_type, chunk_count, processing_status")\
            .eq("client_id", client_id)\
            .eq("processing_status", "completed")\
            .execute()

        if not docs.data:
            return {
                "success": False,
                "error": "No completed documents found in document_uploads",
                "client_id": client_id
            }

        results = []
        total_embeddings_created = 0
        chunks_sources_checked = []

        for doc in docs.data:
            doc_id = doc["id"]
            filename = doc["filename"]
            chunks_data = []

            try:
                # Try document_chunks first
                try:
                    chunks = supabase.table("document_chunks")\
                        .select("id, chunk_text, chunk_index")\
                        .eq("document_id", doc_id)\
                        .order("chunk_index")\
                        .execute()
                    if chunks.data:
                        chunks_data = chunks.data
                        chunks_sources_checked.append("document_chunks")
                except Exception:
                    pass

                # Try vector_embeddings (already has chunks with embeddings in different format)
                if not chunks_data:
                    try:
                        vec_emb = supabase.table("vector_embeddings")\
                            .select("id, chunk_id, embedding")\
                            .eq("client_id", client_id)\
                            .limit(50)\
                            .execute()
                        if vec_emb.data:
                            # Get chunk text from document_chunks using chunk_id
                            for ve in vec_emb.data:
                                chunk_id = ve.get("chunk_id")
                                if chunk_id:
                                    chunk_data = supabase.table("document_chunks")\
                                        .select("chunk_text, chunk_index, document_id")\
                                        .eq("id", chunk_id)\
                                        .execute()
                                    if chunk_data.data and chunk_data.data[0].get("document_id") == doc_id:
                                        chunks_data.append({
                                            "chunk_text": chunk_data.data[0]["chunk_text"],
                                            "chunk_index": chunk_data.data[0]["chunk_index"],
                                            "existing_embedding": ve.get("embedding")
                                        })
                            if chunks_data:
                                chunks_sources_checked.append("vector_embeddings+document_chunks")
                    except Exception as ve_err:
                        logger.warning(f"vector_embeddings check failed: {ve_err}")

                if not chunks_data:
                    logger.warning(f"No chunks found for document {filename} ({doc_id})")
                    results.append({
                        "document_id": doc_id,
                        "filename": filename,
                        "status": "skipped",
                        "reason": "No chunks found in document_chunks or vector_embeddings"
                    })
                    continue

                embeddings_created = 0

                for chunk in chunks_data:
                    chunk_text = chunk.get("chunk_text", "")
                    chunk_index = chunk.get("chunk_index", 0)

                    if not chunk_text or len(chunk_text) < 10:
                        continue

                    # Check if embedding already exists in document_embeddings
                    existing = supabase.table("document_embeddings")\
                        .select("id")\
                        .eq("document_id", doc_id)\
                        .eq("chunk_index", chunk_index)\
                        .execute()

                    if existing.data:
                        logger.info(f"  Embedding already exists for chunk {chunk_index}")
                        continue

                    # Use existing embedding if available, otherwise generate new one
                    embedding = chunk.get("existing_embedding")
                    if not embedding:
                        try:
                            response = openai_client.embeddings.create(
                                model="text-embedding-ada-002",
                                input=chunk_text[:8000]
                            )
                            embedding = response.data[0].embedding
                        except Exception as emb_error:
                            logger.error(f"Error creating embedding for chunk {chunk_index}: {emb_error}")
                            continue

                    # Insert into document_embeddings
                    embedding_record = {
                        "document_id": doc_id,
                        "client_id": client_id,
                        "chunk_text": chunk_text,
                        "chunk_index": chunk_index,
                        "embedding": embedding,
                        "metadata": {
                            "filename": filename,
                            "char_count": len(chunk_text),
                            "source": "reprocess_endpoint"
                        },
                        "created_at": datetime.utcnow().isoformat()
                    }

                    supabase.table("document_embeddings").insert(embedding_record).execute()
                    embeddings_created += 1
                    total_embeddings_created += 1

                results.append({
                    "document_id": doc_id,
                    "filename": filename,
                    "status": "success",
                    "chunks_found": len(chunks_data),
                    "embeddings_created": embeddings_created
                })

            except Exception as doc_error:
                logger.error(f"Error processing document {filename}: {doc_error}")
                results.append({
                    "document_id": doc_id,
                    "filename": filename,
                    "status": "error",
                    "error": str(doc_error)
                })

        logger.info(f"✅ Reprocessing complete: {total_embeddings_created} embeddings created")

        return {
            "success": True,
            "client_id": client_id,
            "company_name": company_name,
            "documents_processed": len(results),
            "total_embeddings_created": total_embeddings_created,
            "chunks_sources_checked": list(set(chunks_sources_checked)),
            "results": results
        }

    except Exception as e:
        logger.error(f"Reprocess embeddings failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }


@router.post("/populate-embeddings/{client_id}")
async def populate_embeddings_from_chunks(client_id: str):
    """
    SIMPLE endpoint to populate document_embeddings from document_chunks.

    This is the fix for RAG not working - chunks exist but embeddings don't.

    Process:
    1. Read chunks from document_chunks for this client
    2. Generate OpenAI embeddings for each chunk_text
    3. Insert into document_embeddings table (which RAG queries)
    """
    from openai import OpenAI
    import os

    try:
        supabase = get_supabase()
        openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        # Get client info
        client = supabase.table("clients").select("company_name").eq("client_id", client_id).execute()
        company_name = client.data[0].get("company_name") if client.data else "Unknown"

        logger.info(f"🔄 Populating embeddings for {company_name} ({client_id})")

        # Get chunks from document_chunks
        chunks = supabase.table("document_chunks")\
            .select("id, document_id, chunk_text, chunk_index, client_id")\
            .eq("client_id", client_id)\
            .order("chunk_index")\
            .execute()

        if not chunks.data:
            return {
                "success": False,
                "error": "No chunks found in document_chunks table",
                "client_id": client_id
            }

        logger.info(f"📄 Found {len(chunks.data)} chunks in document_chunks")

        # Get document filenames for metadata
        doc_ids = list(set(c["document_id"] for c in chunks.data))
        docs = supabase.table("document_uploads")\
            .select("id, filename")\
            .in_("id", doc_ids)\
            .execute()

        doc_filenames = {d["id"]: d["filename"] for d in (docs.data or [])}

        embeddings_created = 0
        skipped = 0
        errors = []

        for chunk in chunks.data:
            document_id = chunk["document_id"]
            chunk_text = chunk["chunk_text"]
            chunk_index = chunk["chunk_index"]
            filename = doc_filenames.get(document_id, "unknown")

            # Check if embedding already exists
            existing = supabase.table("document_embeddings")\
                .select("id")\
                .eq("document_id", document_id)\
                .eq("chunk_index", chunk_index)\
                .execute()

            if existing.data:
                skipped += 1
                continue

            # Generate embedding
            try:
                response = openai_client.embeddings.create(
                    model="text-embedding-ada-002",
                    input=chunk_text[:8000]
                )
                embedding = response.data[0].embedding

                # Insert into document_embeddings
                embedding_record = {
                    "document_id": document_id,
                    "client_id": client_id,
                    "chunk_text": chunk_text,
                    "chunk_index": chunk_index,
                    "embedding": embedding,
                    "metadata": {
                        "filename": filename,
                        "char_count": len(chunk_text),
                        "source": "populate_embeddings_endpoint"
                    },
                    "created_at": datetime.utcnow().isoformat()
                }

                supabase.table("document_embeddings").insert(embedding_record).execute()
                embeddings_created += 1
                logger.info(f"  ✅ Created embedding for chunk {chunk_index} of {filename}")

            except Exception as e:
                error_msg = f"Chunk {chunk_index}: {str(e)}"
                errors.append(error_msg)
                logger.error(f"  ❌ {error_msg}")

        logger.info(f"✅ Populate embeddings complete: {embeddings_created} created, {skipped} skipped")

        return {
            "success": True,
            "client_id": client_id,
            "company_name": company_name,
            "chunks_found": len(chunks.data),
            "embeddings_created": embeddings_created,
            "skipped_existing": skipped,
            "errors": errors
        }

    except Exception as e:
        logger.error(f"Populate embeddings failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }


@router.post("/sync-embeddings-tables")
async def sync_embeddings_tables():
    """
    One-time sync: Copy embeddings from vector_embeddings to document_embeddings.

    This fixes existing clients where embeddings were created before the dual-write fix.
    The RAG function match_knowledge_embeddings queries document_embeddings,
    but old uploads only wrote to vector_embeddings.

    Process:
    1. Get all rows from vector_embeddings (with chunk data from document_chunks)
    2. For each row, check if it already exists in document_embeddings
    3. If not, insert into document_embeddings
    """
    try:
        supabase = get_supabase()

        logger.info("🔄 Starting embeddings table sync...")

        # Get all vector_embeddings with their chunk data
        vector_embs = supabase.table("vector_embeddings")\
            .select("id, chunk_id, client_id, embedding, document_type, created_at")\
            .execute()

        if not vector_embs.data:
            return {
                "success": True,
                "message": "No embeddings in vector_embeddings to sync",
                "synced": 0
            }

        logger.info(f"Found {len(vector_embs.data)} embeddings in vector_embeddings")

        synced = 0
        skipped = 0
        errors = []

        for ve in vector_embs.data:
            try:
                chunk_id = ve.get("chunk_id")
                client_id = ve.get("client_id")
                embedding = ve.get("embedding")

                if not chunk_id or not embedding:
                    skipped += 1
                    continue

                # Get chunk details
                chunk = supabase.table("document_chunks")\
                    .select("document_id, chunk_text, chunk_index")\
                    .eq("id", chunk_id)\
                    .execute()

                if not chunk.data:
                    skipped += 1
                    continue

                chunk_data = chunk.data[0]
                document_id = chunk_data.get("document_id")
                chunk_text = chunk_data.get("chunk_text")
                chunk_index = chunk_data.get("chunk_index", 0)

                # Check if already exists in document_embeddings
                existing = supabase.table("document_embeddings")\
                    .select("id")\
                    .eq("document_id", document_id)\
                    .eq("chunk_index", chunk_index)\
                    .eq("client_id", client_id)\
                    .execute()

                if existing.data:
                    skipped += 1
                    continue

                # Insert into document_embeddings
                rag_record = {
                    "document_id": document_id,
                    "client_id": client_id,
                    "chunk_text": chunk_text,
                    "chunk_index": chunk_index,
                    "embedding": embedding,
                    "metadata": {
                        "document_type": ve.get("document_type", "unknown"),
                        "char_count": len(chunk_text) if chunk_text else 0,
                        "source": "sync_embeddings_tables"
                    },
                    "created_at": ve.get("created_at") or datetime.utcnow().isoformat()
                }

                supabase.table("document_embeddings").insert(rag_record).execute()
                synced += 1

            except Exception as row_e:
                errors.append(str(row_e))
                if len(errors) > 10:
                    break  # Stop if too many errors

        logger.info(f"✅ Sync complete: {synced} synced, {skipped} skipped, {len(errors)} errors")

        return {
            "success": True,
            "synced": synced,
            "skipped": skipped,
            "errors": errors[:5] if errors else [],
            "total_in_vector_embeddings": len(vector_embs.data)
        }

    except Exception as e:
        logger.error(f"Sync embeddings tables failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }


# =============================================================================
# VOICE PROFILE CRAWLING ENDPOINTS
# =============================================================================


class VoiceCrawlRequest(BaseModel):
    client_id: str
    subreddit: str
    user_limit: int = 100
    comments_per_user: int = 20


@router.post("/crawl-voice-profile")
async def crawl_voice_profile(request: VoiceCrawlRequest, background_tasks: BackgroundTasks):
    """
    Manually trigger voice profile crawl for a subreddit.

    This endpoint crawls top users in the subreddit and extracts their
    writing patterns (avg word count, formality, slang, phrases, etc.)

    Body:
    {
        "client_id": "uuid",
        "subreddit": "HomeImprovement",
        "user_limit": 100,  # Number of users to analyze
        "comments_per_user": 20  # Comments per user to collect
    }

    Returns immediately, runs in background for large crawls.
    For small crawls (user_limit <= 50), runs synchronously.
    """
    import asyncio
    import traceback

    try:
        logger.info(f"🎤 Voice profile crawl requested for r/{request.subreddit}")

        # For small crawls, run synchronously to return results immediately
        if request.user_limit <= 50:
            try:
                from workers.voice_database_worker import VoiceDatabaseWorker

                worker = VoiceDatabaseWorker()

                # Run the async function
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                profile = loop.run_until_complete(
                    worker.analyze_subreddit_voice(
                        subreddit_name=request.subreddit,
                        client_id=request.client_id,
                        user_limit=request.user_limit,
                        comments_per_user=request.comments_per_user
                    )
                )
                loop.close()

                return {
                    "success": True,
                    "mode": "synchronous",
                    "subreddit": request.subreddit,
                    "users_analyzed": profile.get("users_analyzed", 0),
                    "comments_analyzed": profile.get("comments_analyzed", 0),
                    "profile_summary": {
                        "avg_word_count": profile.get("avg_word_count"),
                        "formality_score": profile.get("formality_score"),
                        "capitalization_style": profile.get("capitalization_style"),
                        "common_phrases": profile.get("common_phrases", [])[:5],
                        "slang_examples": profile.get("slang_examples", [])[:5],
                        "emoji_frequency": profile.get("emoji_frequency"),
                        "dominant_tone": profile.get("dominant_tone"),
                        "voice_description": profile.get("voice_description", "")[:200]
                    }
                }

            except Exception as e:
                error_tb = traceback.format_exc()
                logger.error(f"❌ Voice crawl failed: {e}")
                return {
                    "success": False,
                    "error": str(e),
                    "traceback": error_tb
                }

        else:
            # For large crawls, run in background
            async def run_voice_crawl():
                try:
                    from workers.voice_database_worker import VoiceDatabaseWorker

                    worker = VoiceDatabaseWorker()
                    await worker.analyze_subreddit_voice(
                        subreddit_name=request.subreddit,
                        client_id=request.client_id,
                        user_limit=request.user_limit,
                        comments_per_user=request.comments_per_user
                    )
                    logger.info(f"✅ Background voice crawl completed for r/{request.subreddit}")
                except Exception as e:
                    logger.error(f"❌ Background voice crawl failed: {e}")

            background_tasks.add_task(run_voice_crawl)

            return {
                "success": True,
                "mode": "background",
                "message": f"Voice profile crawl started for r/{request.subreddit}",
                "subreddit": request.subreddit,
                "user_limit": request.user_limit,
                "comments_per_user": request.comments_per_user,
                "estimated_time": "2-5 minutes"
            }

    except Exception as e:
        logger.error(f"Error starting voice crawl: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/voice-profiles/{client_id}")
async def get_voice_profiles(client_id: str):
    """
    Get all voice profiles for a client.

    Returns:
        List of voice profiles with key metrics
    """
    try:
        supabase = get_supabase()

        profiles = supabase.table("voice_profiles")\
            .select("*")\
            .eq("client_id", client_id)\
            .execute()

        if not profiles.data:
            return {
                "success": True,
                "client_id": client_id,
                "profiles": [],
                "count": 0,
                "message": "No voice profiles found for this client"
            }

        # Format profiles for response
        formatted = []
        for p in profiles.data:
            vp = p.get("voice_profile", {}) or {}
            formatted.append({
                "subreddit": p.get("subreddit"),
                "users_analyzed": vp.get("users_analyzed", 0) or p.get("users_analyzed", 0),
                "comments_analyzed": vp.get("comments_analyzed", 0) or p.get("comments_analyzed", 0),
                "avg_word_count": vp.get("avg_word_count"),
                "formality_score": vp.get("formality_score") or p.get("formality_score"),
                "capitalization_style": vp.get("capitalization_style"),
                "common_phrases": vp.get("common_phrases", [])[:5],
                "slang_examples": vp.get("slang_examples", [])[:5],
                "emoji_frequency": vp.get("emoji_frequency"),
                "dominant_tone": vp.get("dominant_tone") or p.get("dominant_tone"),
                "last_crawl_date": vp.get("last_crawl_date"),
                "is_fallback": vp.get("is_fallback", False)
            })

        return {
            "success": True,
            "client_id": client_id,
            "profiles": formatted,
            "count": len(formatted)
        }

    except Exception as e:
        logger.error(f"Error fetching voice profiles: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/build-all-voice-profiles/{client_id}")
async def build_all_voice_profiles(client_id: str, background_tasks: BackgroundTasks, user_limit: int = 100, comments_per_user: int = 20):
    """
    Build voice profiles for ALL subreddits configured for a client.

    Runs in background and builds profiles for each subreddit in
    the client's client_subreddit_config table.
    """
    try:
        supabase = get_supabase()

        # Get client info
        client = supabase.table("clients").select("company_name").eq("client_id", client_id).execute()
        if not client.data:
            raise HTTPException(status_code=404, detail="Client not found")

        company_name = client.data[0].get("company_name")

        # Get configured subreddits
        subs = supabase.table("client_subreddit_config")\
            .select("subreddit_name")\
            .eq("client_id", client_id)\
            .eq("is_active", True)\
            .execute()

        subreddits = [s["subreddit_name"] for s in (subs.data or [])]

        if not subreddits:
            # Fallback to subreddits from opportunities
            opps = supabase.table("opportunities")\
                .select("subreddit")\
                .eq("client_id", client_id)\
                .execute()
            if opps.data:
                subreddits = list(set([o["subreddit"] for o in opps.data if o.get("subreddit")]))[:10]

        if not subreddits:
            return {
                "success": False,
                "error": "No subreddits configured for this client"
            }

        logger.info(f"🎤 Building voice profiles for {len(subreddits)} subreddits for {company_name}")

        # Run in background
        async def build_profiles():
            try:
                from workers.voice_database_worker import build_client_voice_database

                result = await build_client_voice_database(
                    client_id=client_id,
                    user_limit=user_limit,
                    comments_per_user=comments_per_user
                )
                logger.info(f"✅ Voice profiles built for {company_name}: {result}")
            except Exception as e:
                logger.error(f"❌ Error building voice profiles: {e}")

        background_tasks.add_task(build_profiles)

        return {
            "success": True,
            "message": f"Building voice profiles for {len(subreddits)} subreddits",
            "client_id": client_id,
            "company_name": company_name,
            "subreddits": subreddits,
            "user_limit": user_limit,
            "comments_per_user": comments_per_user,
            "estimated_time": f"{len(subreddits) * 2}-{len(subreddits) * 5} minutes"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error building voice profiles: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/voice-profile-freshness/{client_id}")
async def get_voice_profile_freshness(client_id: str, max_age_days: int = 30):
    """
    Check the freshness of voice profiles for a client.

    Voice profiles should be refreshed every 30 days to stay current with
    evolving community language patterns.

    Args:
        client_id: Client UUID
        max_age_days: Maximum age in days before profile is considered stale (default 30)

    Returns:
        Freshness status for each profile, including which need refresh
    """
    try:
        from workers.voice_database_worker import check_voice_profile_freshness

        result = check_voice_profile_freshness(client_id, max_age_days)

        return {
            "success": True,
            **result
        }

    except Exception as e:
        logger.error(f"Error checking voice profile freshness: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/trigger-voice-refresh")
async def trigger_voice_profile_refresh(
    background_tasks: BackgroundTasks,
    client_id: str = None,
    max_age_days: int = 30,
    user_limit: int = 100,
    comments_per_user: int = 20
):
    """
    Trigger refresh of stale voice profiles.

    This endpoint refreshes all voice profiles older than max_age_days.
    Can be called manually or set up as a scheduled job.

    Args:
        client_id: Optional - refresh for specific client, or all clients if None
        max_age_days: Profiles older than this will be refreshed (default 30)
        user_limit: Users to analyze per subreddit (default 100)
        comments_per_user: Comments per user (default 20)

    Returns immediately, refresh runs in background.
    """
    try:
        logger.info(f"🔄 Voice profile refresh triggered (client: {client_id or 'ALL'}, max_age: {max_age_days} days)")

        async def run_refresh():
            try:
                from workers.voice_database_worker import refresh_stale_voice_profiles

                result = await refresh_stale_voice_profiles(
                    client_id=client_id,
                    max_age_days=max_age_days,
                    user_limit=user_limit,
                    comments_per_user=comments_per_user
                )
                logger.info(f"✅ Voice refresh complete: {result.get('refreshed', 0)} refreshed, {result.get('failed', 0)} failed")
            except Exception as e:
                logger.error(f"❌ Voice refresh failed: {e}")

        background_tasks.add_task(run_refresh)

        return {
            "success": True,
            "message": "Voice profile refresh started in background",
            "client_id": client_id or "ALL",
            "max_age_days": max_age_days,
            "user_limit": user_limit,
            "comments_per_user": comments_per_user,
            "note": "Check logs or call GET /voice-profile-freshness/{client_id} to verify"
        }

    except Exception as e:
        logger.error(f"Error triggering voice refresh: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/generate-content-excel/{client_id}")
async def generate_content_excel(
    client_id: str,
    limit: int = 25
):
    """
    Generate content for opportunities and export directly to Excel.

    This endpoint:
    1. Gets top opportunities for the client
    2. Generates content using anti-AI voice matching
    3. Exports directly to Excel with 31 columns matching RECHO format

    Args:
        client_id: Client UUID
        limit: Number of opportunities to process (default 25, max 50)

    Returns:
        Excel file download with voice-matched content
    """
    import io
    import json as json_lib
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    try:
        supabase = get_supabase()
        limit = min(limit, 50)  # Cap at 50

        logger.info(f"📊 Generating content Excel for client {client_id} (limit: {limit})")

        # Get client info
        client_response = supabase.table("clients").select("*").eq("client_id", client_id).execute()
        if not client_response.data:
            raise HTTPException(status_code=404, detail="Client not found")

        client = client_response.data[0]
        company_name = client.get("company_name", "Client")
        industry = client.get("industry", "").lower()

        # Check if medical disclaimer needed based on industry
        medical_industries = ['health', 'medical', 'wellness', 'supplement', 'vitamin', 'pharmaceutical', 'fertility']
        needs_medical_disclaimer = any(kw in industry for kw in medical_industries)

        # Get opportunities with scoring data - sort by weighted impact (highest first)
        opps_response = supabase.table("opportunities")            .select("*")            .eq("client_id", client_id)            .limit(limit * 2)            .execute()

        if not opps_response.data:
            raise HTTPException(status_code=404, detail="No opportunities found for this client")

        # Calculate weighted scores and sort by impact (highest first)
        def calculate_weighted_score(opp):
            commercial = opp.get('commercial_intent_score', 0) or 0
            relevance = opp.get('relevance_score', 0) or 0
            engagement = opp.get('engagement_score', 0) or 0
            timing = opp.get('timing_score', 0) or 0
            return commercial * 0.35 + relevance * 0.25 + engagement * 0.20 + timing * 0.20

        opportunities = sorted(opps_response.data, key=calculate_weighted_score, reverse=True)[:limit]
        logger.info(f"📊 Found {len(opportunities)} opportunities (sorted by weighted impact)")

        # Build opportunity lookup map for scoring data
        opp_lookup = {opp.get('opportunity_id'): opp for opp in opportunities}

        # Generate content using ContentGenerationWorker
        from workers.content_generation_worker import ContentGenerationWorker
        worker = ContentGenerationWorker()

        result = worker.generate_content_for_client(
            client_id=client_id,
            opportunities=opportunities,
            delivery_batch=f"EXCEL-{datetime.now().strftime('%Y-%m-%d')}"
        )

        if not result.get("success"):
            raise HTTPException(status_code=500, detail=f"Content generation failed: {result.get('error')}")

        content_items = result.get("content", [])
        logger.info(f"📊 Generated {len(content_items)} content items")

        # Get voice profiles for enhanced proof
        voice_profiles = {}
        try:
            voice_response = supabase.table("voice_profiles").select("*").eq("client_id", client_id).execute()
            for vp in (voice_response.data or []):
                subreddit = vp.get('subreddit_name', '').lower()
                voice_profiles[subreddit] = vp
        except Exception as e:
            logger.warning(f"Could not load voice profiles: {e}")


        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Organic Content"

        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # 31 Column Headers (matching RECHO format)
        headers = [
            "Opportunity ID",           # 1
            "Date Found",               # 2
            "Subreddit",                # 3
            "Thread Title",             # 4
            "Thread URL",               # 5
            "Original Post/Comment",    # 6
            "Context Summary",          # 7 (AI-generated)
            "Commercial Intent Score",  # 8
            "Relevance Score",          # 9
            "Engagement Score",         # 10
            "Timing Score",             # 11
            "Overall Priority",         # 12 (High/Medium/Low)
            "Urgency Level",            # 13
            "Buying Signal Location",   # 14 (AI-generated)
            "Content Type",             # 15
            "Suggested Reply/Post",     # 16 (THE CONTENT)
            "Revised",                  # 17 (empty for manual edits)
            "Voice Similarity Proof",   # 18
            "Tone Match",               # 19 (AI-generated)
            "Product Link",             # 20 (N/A if no brand)
            "Medical Disclaimer Needed?", # 21
            "Ideal Engagement Window",  # 22
            "Mod-Friendly?",            # 23
            "Posting Window",           # 24
            "Follow-up Timing",         # 25
            "Follow-up Strategy",       # 26 (AI-generated)
            "Backup Plan",              # 27 (AI-generated)
            "Thread Status",            # 28
            "Keywords/Tags",            # 29
            "Additional Notes",         # 30
            "Posting Account",          # 31
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Column widths
        column_widths = [
            15, 18, 14, 50, 50, 60, 50, 12, 12, 12, 12, 14, 12, 40, 14,
            120, 60, 50, 30, 30, 12, 20, 12, 20, 20, 40, 40, 14, 30, 40, 18
        ]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        # Write content rows
        for row_idx, item in enumerate(content_items, 2):
            opp_id = item.get('opportunity_id')
            opp = opp_lookup.get(opp_id, {})

            # Parse matched keywords
            matched_keywords = item.get('matched_keywords', '') or opp.get('matched_keywords', '')
            if isinstance(matched_keywords, list):
                matched_keywords = ', '.join(matched_keywords)
            elif matched_keywords and matched_keywords.startswith('['):
                try:
                    matched_keywords = ', '.join(json_lib.loads(matched_keywords))
                except:
                    pass

            # GET REAL SCORES from opportunity data (0-10 scale)
            commercial_intent = opp.get('commercial_intent_score', 0) or 0
            relevance = opp.get('relevance_score', 0) or 0
            engagement = opp.get('engagement_score', 0) or 0
            timing = opp.get('timing_score', 0) or 0

            # Calculate overall priority using weighted formula
            weighted_score = commercial_intent * 0.35 + relevance * 0.25 + engagement * 0.20 + timing * 0.20
            if weighted_score >= 7.5:
                overall_priority = "High"
            elif weighted_score >= 5:
                overall_priority = "Medium"
            else:
                overall_priority = "Low"

            # Urgency level based on timing + commercial intent
            if timing > 8 and commercial_intent > 7:
                urgency_level = "Urgent"
            elif timing > 7 or commercial_intent > 7:
                urgency_level = "High"
            else:
                urgency_level = "Medium"

            # Get original text and thread title
            original_text = (item.get('original_post_text', '') or opp.get('original_post_text', '') or '')[:500]
            thread_title = item.get('thread_title', '') or opp.get('thread_title', '')

            # Context Summary (AI-generated)
            context_summary = f"User seeking advice about {thread_title[:50]}. Shows interest in community recommendations."
            if 'recommend' in original_text.lower() or 'suggestion' in original_text.lower():
                context_summary = f"Active recommendation request: {thread_title[:40]}. High engagement potential."
            elif 'help' in original_text.lower() or 'question' in original_text.lower():
                context_summary = f"Help-seeking post about {thread_title[:40]}. Good opportunity for expert advice."

            # EXTRACT ACTUAL BUYING SIGNAL QUOTES
            import re as re_local
            buying_keywords = ['buy', 'purchase', 'looking for', 'recommend', 'best', 'worth', 'price', 'budget', 'cost', 'afford', 'get one', 'should i']
            sentences = re_local.split(r'[.!?\n]+', original_text) if original_text else []
            found_quotes = []
            for sentence in sentences:
                sentence_lower = sentence.lower().strip()
                for keyword in buying_keywords:
                    if keyword in sentence_lower and len(sentence.strip()) > 15:
                        quote = sentence.strip()[:120]
                        if quote and quote not in [q.strip('"') for q in found_quotes]:
                            found_quotes.append(f'"{quote}"')
                        break
            buying_signal_location = " | ".join(found_quotes[:2]) if found_quotes else "No explicit buying signals - general discussion"

            # Tone Match and formality
            tone = item.get('tone', 'conversational')
            formality = item.get('formality_score', 0.5)
            subreddit = item.get('subreddit', '') or opp.get('subreddit', '')

            # BUILD ENHANCED VOICE SIMILARITY PROOF
            sub_lower = subreddit.lower().replace('r/', '')
            vp = voice_profiles.get(sub_lower, {})
            proof_parts = []
            vp_formality = vp.get('formality_score') or formality
            vp_tone = vp.get('dominant_tone') or tone
            proof_parts.append(f"r/{subreddit} voice: formality={round(vp_formality, 2)}, tone={vp_tone}")
            content_text = item.get('text', '')
            word_count = len(content_text.split()) if content_text else 0
            avg_words = vp.get('avg_word_count', 75)
            proof_parts.append(f"Length ({word_count} words) vs avg ({avg_words})")
            style_notes = []
            if content_text and '!' not in content_text[:200]:
                style_notes.append("no exclamation marks")
            if content_text and not any(e in content_text for e in ['\U0001f44d', '\U0001f60a', '\U0001f525', '\U0001f4af']):
                style_notes.append("no emojis")
            if formality > 0.6:
                style_notes.append("proper capitalization")
            if style_notes:
                proof_parts.append(f"Style: {', '.join(style_notes)}")
            common_vocab = vp.get('common_vocabulary', [])
            if common_vocab and isinstance(common_vocab, list):
                proof_parts.append(f"Vocab: {', '.join(str(v) for v in common_vocab[:3])}")
            voice_similarity_proof = " | ".join(proof_parts)

            # Tone match description
            if formality < 0.3:
                tone_match = f"Casual/informal - matches r/{subreddit} community voice"
            elif formality < 0.6:
                tone_match = f"Conversational - {tone}"
            else:
                tone_match = f"Semi-formal/professional - {tone}"

            # PRODUCT LINK - Only show if brand ACTUALLY appears in generated content
            brand_in_content = company_name.lower() in content_text.lower() if content_text and company_name else False
            if brand_in_content:
                product_link = client.get('website_url', 'See brand website')
            else:
                product_link = "N/A"

            # Ideal Engagement Window
            posting_time = datetime.now()
            ideal_window = f"{posting_time.strftime('%I:%M %p')} - {(posting_time.replace(hour=(posting_time.hour + 4) % 24)).strftime('%I:%M %p')} EST"

            # Follow-up Strategy
            follow_up_strategy = "Monitor for replies within 24h. If positive engagement, provide additional value. If questions arise, respond helpfully without being pushy."

            # Backup Plan
            backup_plan = "If thread gets locked or removed, save content for similar future opportunities. Consider posting in related subreddits if appropriate."

            # Thread Status
            thread_status = "Active" if timing >= 5 else "Aging"

            # Row data (31 columns)
            row_data = [
                str(opp_id or '')[:36],                              # 1: Opportunity ID
                item.get('date_found', '') or opp.get('date_found', ''),  # 2: Date Found
                f"r/{subreddit}",                                    # 3: Subreddit
                thread_title,                                        # 4: Thread Title
                item.get('thread_url', '') or opp.get('thread_url', ''),  # 5: Thread URL
                original_text,                                       # 6: Original Post/Comment
                context_summary,                                     # 7: Context Summary (AI)
                commercial_intent,                                   # 8: Commercial Intent Score (0-10)
                relevance,                                           # 9: Relevance Score (0-10)
                engagement,                                          # 10: Engagement Score (0-10)
                timing,                                              # 11: Timing Score (0-10)
                overall_priority,                                    # 12: Overall Priority
                urgency_level,                                       # 13: Urgency Level
                buying_signal_location,                              # 14: Buying Signal Location (actual quotes)
                item.get('type', 'REPLY').upper(),                  # 15: Content Type
                content_text,                                        # 16: Suggested Reply/Post
                "",                                                  # 17: Revised (empty)
                voice_similarity_proof,                              # 18: Voice Similarity Proof (enhanced)
                tone_match,                                          # 19: Tone Match (AI)
                product_link,                                        # 20: Product Link (N/A if no brand in content)
                "Yes" if needs_medical_disclaimer else "No",        # 21: Medical Disclaimer Needed?
                ideal_window,                                        # 22: Ideal Engagement Window
                "Yes",                                               # 23: Mod-Friendly?
                "Within 4 hours",                                    # 24: Posting Window
                "24-48 hours",                                       # 25: Follow-up Timing
                follow_up_strategy,                                  # 26: Follow-up Strategy (AI)
                backup_plan,                                         # 27: Backup Plan (AI)
                thread_status,                                       # 28: Thread Status
                matched_keywords,                                    # 29: Keywords/Tags
                f"Voice: {tone} | KB: {item.get('knowledge_insights_used', 0)} insights | Score: {round(weighted_score, 2)}",  # 30: Additional Notes
                item.get('assigned_profile', 'TBD'),                # 31: Posting Account
            ]

            # Write row
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border

                # Wrap text for content columns
                if col_idx in [6, 7, 14, 16, 17, 18, 19, 26, 27, 30]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                elif col_idx == 12:  # Overall Priority colors
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if value == "High":
                        cell.font = Font(color="FF0000", bold=True)
                    elif value == "Medium":
                        cell.font = Font(color="FFA500", bold=True)
                    else:
                        cell.font = Font(color="008000", bold=True)
                else:
                    cell.alignment = Alignment(vertical='center')

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"{company_name.replace(' ', '_')}_RECHO_Weekly_Organic_Content_{today}.xlsx"

        logger.info(f"✅ Generated Excel: {filename} with {len(content_items)} rows")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"❌ Excel generation failed: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/admin/debug-scored-opportunities/{client_id}")
async def debug_scored_opportunities(client_id: str, limit: int = 5):
    """Debug endpoint to check if opportunity scores are being saved"""
    supabase = get_supabase_client()

    # Get opportunities with scoring columns
    response = supabase.table("opportunities")\
        .select("opportunity_id, commercial_intent_score, relevance_score, engagement_score, composite_score, priority_tier, updated_at")\
        .eq("client_id", client_id)\
        .not_.is_("composite_score", "null")\
        .order("composite_score", desc=True)\
        .limit(limit)\
        .execute()

    scored_count = supabase.table("opportunities")\
        .select("opportunity_id", count="exact")\
        .eq("client_id", client_id)\
        .not_.is_("composite_score", "null")\
        .execute()

    unscored_count = supabase.table("opportunities")\
        .select("opportunity_id", count="exact")\
        .eq("client_id", client_id)\
        .is_("composite_score", "null")\
        .execute()

    return {
        "client_id": client_id,
        "scored_opportunities": scored_count.count if hasattr(scored_count, 'count') else len(scored_count.data or []),
        "unscored_opportunities": unscored_count.count if hasattr(unscored_count, 'count') else len(unscored_count.data or []),
        "top_scored": response.data
    }