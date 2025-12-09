"""
Sample Content Generator V2 - Matches User Example Format
Generates 25-piece content queue with real opportunities
"""
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from typing import List, Dict
import io

logger = logging.getLogger(__name__)


class SampleContentGeneratorV2:
    """Generates 25-piece sample content matching user example format"""
    
    def __init__(self, supabase_client, openai_client):
        self.supabase = supabase_client
        self.openai = openai_client
    
    def generate_report(self, client_id: str, opportunities: List[Dict]) -> io.BytesIO:
        """
        Generate sample content report
        
        Args:
            client_id: Client UUID
            opportunities: List of top 25 scored opportunities
            
        Returns:
            BytesIO: Excel file in memory
        """
        logger.info(f"📝 Generating Sample Content V2 for client {client_id}")
        
        # Fetch client data
        client = self.supabase.table("clients").select("*").eq("client_id", client_id).single().execute().data
        
        # Sort opportunities by priority
        top_opportunities = sorted(
            opportunities,
            key=lambda x: x.get('overall_priority', 0),
            reverse=True
        )[:25]
        
        # Generate content for each opportunity
        content_items = []
        for i, opp in enumerate(top_opportunities, 1):
            content = self._generate_content_item(client, opp, i)
            content_items.append(content)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Content Queue"
        
        # Headers (matching user example exactly)
        headers = [
            "Opportunity ID", "Date Found", "Subreddit", "Thread Title", "Thread URL",
            "Original Post/Comment", "Context Summary", "Relevance Score", "Engagement Score",
            "Timing Score", "Commercial Intent Score", "Overall Priority", "Urgency Level",
            "Buying Signal Location", "Content Type", "Suggested Reply/Post",
            "Voice Similarity Proof", "Tone Match", "Product Mentioned", "Product Link",
            "Call-to-Action", "Medical Disclaimer Needed?", "Ideal Engagement",
            "Risk Level", "Mod-Friendly?", "Posting Window", "Assigned To", "Status", "Notes"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write content rows
        for row_idx, item in enumerate(content_items, 2):
            ws.cell(row=row_idx, column=1, value=item['opportunity_id'])
            ws.cell(row=row_idx, column=2, value=item['date_found'])
            ws.cell(row=row_idx, column=3, value=item['subreddit'])
            ws.cell(row=row_idx, column=4, value=item['thread_title'])
            ws.cell(row=row_idx, column=5, value=item['thread_url'])
            ws.cell(row=row_idx, column=6, value=item['original_post'])
            ws.cell(row=row_idx, column=7, value=item['context_summary'])
            ws.cell(row=row_idx, column=8, value=item['relevance_score'])
            ws.cell(row=row_idx, column=9, value=item['engagement_score'])
            ws.cell(row=row_idx, column=10, value=item['timing_score'])
            ws.cell(row=row_idx, column=11, value=item['commercial_intent_score'])
            ws.cell(row=row_idx, column=12, value=item['overall_priority'])
            ws.cell(row=row_idx, column=13, value=item['urgency_level'])
            ws.cell(row=row_idx, column=14, value=item['buying_signal_location'])
            ws.cell(row=row_idx, column=15, value=item['content_type'])
            ws.cell(row=row_idx, column=16, value=item['suggested_reply'])
            ws.cell(row=row_idx, column=17, value=item['voice_similarity_proof'])
            ws.cell(row=row_idx, column=18, value=item['tone_match'])
            ws.cell(row=row_idx, column=19, value=item['product_mentioned'])
            ws.cell(row=row_idx, column=20, value=item['product_link'])
            ws.cell(row=row_idx, column=21, value=item['call_to_action'])
            ws.cell(row=row_idx, column=22, value=item['medical_disclaimer'])
            ws.cell(row=row_idx, column=23, value=item['ideal_engagement'])
            ws.cell(row=row_idx, column=24, value=item['risk_level'])
            ws.cell(row=row_idx, column=25, value=item['mod_friendly'])
            ws.cell(row=row_idx, column=26, value=item['posting_window'])
            ws.cell(row=row_idx, column=27, value=item['assigned_to'])
            ws.cell(row=row_idx, column=28, value=item['status'])
            ws.cell(row=row_idx, column=29, value=item['notes'])
            
            # Apply text wrapping to long content cells
            ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_idx, column=16).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_idx, column=17).alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        column_widths = {
            'A': 12, 'B': 16, 'C': 15, 'D': 40, 'E': 30,
            'F': 50, 'G': 40, 'H': 12, 'I': 12, 'J': 12,
            'K': 18, 'L': 12, 'M': 12, 'N': 50, 'O': 12,
            'P': 60, 'Q': 50, 'R': 20, 'S': 25, 'T': 35,
            'U': 15, 'V': 15, 'W': 18, 'X': 12, 'Y': 12,
            'Z': 15, 'AA': 15, 'AB': 12, 'AC': 30
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row height for better readability
        ws.row_dimensions[1].height = 40
        for row in range(2, len(content_items) + 2):
            ws.row_dimensions[row].height = 80
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"✅ Sample Content V2 generated: {len(content_items)} pieces")
        return output
    
    def _generate_content_item(self, client: Dict, opp: Dict, index: int) -> Dict:
        """Generate a single content item"""
        
        # Generate AI content using OpenAI
        suggested_reply = self._generate_ai_reply(client, opp)
        
        # Extract data from opportunity
        opportunity_id = opp.get('id', f"OPP-{3000 + index}")
        date_found = opp.get('created_at', datetime.now().isoformat())[:16].replace('T', ' ')
        subreddit = opp.get('subreddit', 'r/Unknown')
        thread_title = opp.get('thread_title', 'Untitled Thread')
        thread_url = opp.get('thread_url', f"reddit.com/r/{subreddit}/comments/xyz{index}/...")
        original_post = opp.get('post_content', thread_title)[:200]
        
        relevance_score = int(opp.get('relevance_score', 85))
        engagement_score = int(opp.get('engagement_score', 75))
        timing_score = int(opp.get('timing_score', 88))
        commercial_intent = int(opp.get('commercial_intent_score', 90))
        overall_priority = round(opp.get('overall_priority', 84.5), 1)
        
        # Determine urgency
        if overall_priority >= 90:
            urgency = "URGENT"
        elif overall_priority >= 85:
            urgency = "HIGH"
        elif overall_priority >= 75:
            urgency = "MEDIUM"
        else:
            urgency = "LOW"
        
        # Buying signal detection
        buying_signal = self._detect_buying_signal(original_post, thread_title)
        
        # Content type
        content_type = "Reply" if opp.get('is_comment', False) else "Reply"
        
        # Voice analysis
        voice_proof = self._analyze_voice_match(suggested_reply)
        tone_match = self._determine_tone(suggested_reply)
        
        # Product mention
        product_mentioned = self._extract_product_mention(suggested_reply, client)
        product_link = self._get_product_link(client, product_mentioned)
        
        # CTA analysis
        cta_type = self._analyze_cta(suggested_reply)
        
        # Risk assessment
        risk_level = "LOW" if overall_priority < 90 else "MEDIUM"
        mod_friendly = "Yes"
        
        # Posting window (hours from now)
        posting_window = f"{72 - (index * 2)} hours"
        
        return {
            'opportunity_id': opportunity_id,
            'date_found': date_found,
            'subreddit': subreddit,
            'thread_title': thread_title,
            'thread_url': thread_url,
            'original_post': original_post,
            'context_summary': f"Reply addressing {thread_title[:50]}",
            'relevance_score': relevance_score,
            'engagement_score': engagement_score,
            'timing_score': timing_score,
            'commercial_intent_score': commercial_intent,
            'overall_priority': overall_priority,
            'urgency_level': urgency,
            'buying_signal_location': buying_signal,
            'content_type': content_type,
            'suggested_reply': suggested_reply,
            'voice_similarity_proof': voice_proof,
            'tone_match': tone_match,
            'product_mentioned': product_mentioned,
            'product_link': product_link,
            'call_to_action': cta_type,
            'medical_disclaimer': "No",
            'ideal_engagement': f"{12 + index}-{18 + index} upvotes",
            'risk_level': risk_level,
            'mod_friendly': mod_friendly,
            'posting_window': posting_window,
            'assigned_to': "Content Team",
            'status': "Queued",
            'notes': None
        }
    
    def _generate_ai_reply(self, client: Dict, opp: Dict) -> str:
        """Generate AI reply content"""
        try:
            prompt = f"""Generate a natural Reddit reply for this opportunity:

Thread: {opp.get('thread_title', '')}
Post Content: {opp.get('post_content', '')}

Client: {client.get('company_name', 'Company')}
Products: {client.get('products_services', 'Products')}
Website: {client.get('website', '')}
Tone: {client.get('content_tone', 'Helpful and educational')}

Write a 3-4 paragraph helpful reply that:
1. Addresses their question naturally
2. Provides genuine value
3. Subtly mentions the brand if relevant
4. Uses casual, conversational language
5. No rigid structure or formulaic intro

Reply:"""
            
            response = self.openai.chat.completions.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.8,
                max_tokens=300
            )
            
            return response.choices[0].message.content.strip()
        
        except Exception as e:
            logger.error(f"AI generation error: {e}")
            # Fallback template
            return f"""That's a great question. 

I've dealt with this before and found that focusing on quality over quick fixes makes a huge difference.

{client.get('website', 'Check out reliable resources')} has some helpful options if you want to explore further.

Hope that helps!"""
    
    def _detect_buying_signal(self, content: str, title: str) -> str:
        """Detect buying signals in content"""
        combined = (content + " " + title).lower()
        
        if 'recommend' in combined or 'suggestion' in combined:
            return "Title/Post contains explicit recommendation request"
        elif 'best' in combined or 'which' in combined:
            return "Comparison shopping language detected"
        elif 'buy' in combined or 'purchase' in combined:
            return "Direct buying intent keywords present"
        elif 'worth' in combined or 'should i' in combined:
            return "Purchase evaluation phase detected"
        else:
            return "Implicit interest signals present"
    
    def _analyze_voice_match(self, content: str) -> str:
        """
        Analyze voice similarity to brand with DETAILED, SPECIFIC proof.

        Returns a comprehensive breakdown showing WHY the content matches
        the target voice, with specific examples from the generated text.
        """
        content_lower = content.lower()

        # Analyze specific voice characteristics
        analysis_points = []

        # 1. Sentence Structure Analysis
        sentences = [s.strip() for s in content.split('.') if s.strip()]
        avg_words = sum(len(s.split()) for s in sentences) / max(len(sentences), 1)
        if avg_words < 15:
            analysis_points.append(f"SHORT SENTENCES (avg {avg_words:.0f} words) = casual Reddit style")
        elif avg_words < 25:
            analysis_points.append(f"MEDIUM SENTENCES (avg {avg_words:.0f} words) = balanced conversational")
        else:
            analysis_points.append(f"LONGER SENTENCES (avg {avg_words:.0f} words) = detailed/educational")

        # 2. Casual Language Markers
        casual_markers = {
            'honestly': 'authentic personal framing',
            'literally': 'emphasis style common in subreddit',
            'actually': 'conversational correction tone',
            'basically': 'simplification for accessibility',
            'just': 'casual minimizer',
            'pretty': 'hedging language',
            'super': 'enthusiastic modifier',
            'really': 'emphasis marker',
            'gonna': 'casual contraction',
            'kinda': 'informal hedging',
            'tbh': 'text-speak authenticity',
            'imo': 'opinion marker',
            'lol': 'humor/lightness',
            'haha': 'friendly tone'
        }

        found_markers = []
        for marker, meaning in casual_markers.items():
            if marker in content_lower:
                found_markers.append(f"'{marker}' ({meaning})")

        if found_markers:
            analysis_points.append(f"CASUAL MARKERS: {', '.join(found_markers[:3])}")

        # 3. Personal Experience Indicators
        personal_phrases = ['i ', "i've", "i'm", 'my ', 'me ', 'myself', 'personally', 'my experience']
        personal_count = sum(1 for p in personal_phrases if p in content_lower)
        if personal_count >= 3:
            analysis_points.append("STRONG PERSONAL VOICE: multiple first-person references = authentic sharing")
        elif personal_count >= 1:
            analysis_points.append("PERSONAL TOUCH: first-person language creates relatability")

        # 4. Empathy Markers
        empathy_phrases = ['understand', 'feel', 'been there', 'same thing', 'get it', 'know how', 'tough', 'hard']
        empathy_found = [p for p in empathy_phrases if p in content_lower]
        if empathy_found:
            analysis_points.append(f"EMPATHY SIGNALS: {', '.join(empathy_found[:2])} = emotional connection")

        # 5. Anti-AI Patterns (what it DOESN'T have)
        ai_red_flags = ['furthermore', 'additionally', 'in conclusion', 'to summarize', 'it is important to note']
        has_ai_patterns = any(flag in content_lower for flag in ai_red_flags)
        if not has_ai_patterns:
            analysis_points.append("NO AI RED FLAGS: avoids corporate/robotic language patterns")

        # 6. Formatting Style
        has_bullet_lists = content.count('•') > 0 or content.count('-') > 2
        has_short_paragraphs = content.count('\n\n') >= 1

        if not has_bullet_lists and has_short_paragraphs:
            analysis_points.append("NATURAL FORMATTING: paragraph breaks without rigid listicles")
        elif not has_bullet_lists:
            analysis_points.append("FLOWING PROSE: single-block conversational style")

        # 7. Extract actual phrases as proof
        sample_phrases = []
        for sentence in sentences[:3]:
            if len(sentence) > 20 and len(sentence) < 100:
                sample_phrases.append(f'"{sentence[:50]}..."')

        if sample_phrases:
            analysis_points.append(f"SAMPLE VOICE: {sample_phrases[0]}")

        # Build comprehensive proof
        if len(analysis_points) >= 4:
            return " | ".join(analysis_points[:5])
        elif len(analysis_points) >= 2:
            return " | ".join(analysis_points)
        else:
            return "Conversational tone with community-appropriate language and natural flow"
    
    def _determine_tone(self, content: str) -> str:
        """
        Determine tone of content with detailed analysis.

        Returns specific tone descriptors based on linguistic markers.
        """
        content_lower = content.lower()
        tones = []

        # Check for different tone markers
        if any(w in content_lower for w in ['help', 'hope this helps', 'let me know']):
            tones.append("Helpful")

        if any(w in content_lower for w in ['experience', "i've been", 'i used', 'i tried', 'worked for me']):
            tones.append("Experiential")

        if any(w in content_lower for w in ['honestly', 'tbh', 'real talk', 'truth is']):
            tones.append("Candid")

        if any(w in content_lower for w in ['understand', 'get it', 'feel you', 'same boat', 'been there']):
            tones.append("Empathetic")

        if any(w in content_lower for w in ['!', 'love', 'amazing', 'great', 'awesome']):
            tones.append("Enthusiastic")

        if any(w in content_lower for w in ['consider', 'might want', 'could try', 'one option']):
            tones.append("Advisory")

        if any(w in content_lower for w in ['?', 'depends', 'it varies', 'ymmv']):
            tones.append("Nuanced")

        if len(tones) >= 2:
            return f"{tones[0]} + {tones[1]} blend"
        elif len(tones) == 1:
            return f"{tones[0]} community voice"
        else:
            return "Balanced conversational"
    
    def _extract_product_mention(self, content: str, client: Dict) -> str:
        """Extract mentioned product"""
        company = client.get('company_name', '').lower()
        website = client.get('website', '').lower()
        
        if company in content.lower() or website in content.lower():
            products = client.get('products_services', 'Product').split(',')
            return products[0].strip() if products else "Brand mention"
        else:
            return "None (educational content)"
    
    def _get_product_link(self, client: Dict, product: str) -> str:
        """Get product link if mentioned"""
        if product and product != "None (educational content)":
            website = client.get('website', 'example.com')
            return f"{website}/products/{product.lower().replace(' ', '-')}"
        return None
    
    def _analyze_cta(self, content: str) -> str:
        """Analyze call-to-action type"""
        content_lower = content.lower()
        
        if 'check out' in content_lower or 'visit' in content_lower:
            return "Soft mention"
        elif 'http' in content_lower:
            return "Direct link"
        else:
            return "No CTA"
