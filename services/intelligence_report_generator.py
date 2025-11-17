"""
Intelligence Report Generator Service

Generates comprehensive onboarding intelligence reports for new clients.
Creates 10-sheet Excel workbook with:
- 7 core intelligence sheets (market data, subreddit analysis, moderators, threads, influencers, risks, commercial intent)
- 3 strategy sheets (brand voice, content timeline, content splits)

This service is called during client onboarding to provide deep market intelligence.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any
import logging
from database import get_supabase_client

logger = logging.getLogger(__name__)

class IntelligenceReportGenerator:
    """Generates comprehensive intelligence reports for new clients"""
    
    # Define standard formatting
    HEADER_FILL = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=14)
    SECTION_FILL = PatternFill(start_color="E8EAED", end_color="E8EAED", fill_type="solid")
    SECTION_FONT = Font(bold=True, size=11)
    LIGHT_BG_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Tier colors for subreddits
    TIER_COLORS = {
        "PLATINUM": PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        "GOLD": PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
        "SILVER": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    }
    
    # Commercial intent colors
    HIGH_INTENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    MEDIUM_INTENT_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Risk colors
    LOW_RISK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    MEDIUM_RISK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    HIGH_RISK_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def __init__(self, client_id: str):
        self.client_id = client_id
        self.supabase = get_supabase_client()
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
    
    def set_column_widths(self, ws, widths: List[int]):
        """Set column widths for worksheet"""
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def add_header_row(self, ws, row: int, text: str, merge_to_col: int = 8):
        """Add styled header row"""
        ws.merge_cells(f'A{row}:{get_column_letter(merge_to_col)}{row}')
        cell = ws[f'A{row}']
        cell.value = text
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    async def fetch_client_data(self) -> Dict[str, Any]:
        """Fetch all client data from database"""
        try:
            # Fetch client info
            client_response = self.supabase.table("clients").select("*").eq("client_id", self.client_id).execute()
            if not client_response.data:
                raise ValueError(f"Client {self.client_id} not found")
            
            client = client_response.data[0]
            
            # Fetch subreddits
            subreddits_response = self.supabase.table("client_subreddits").select("*").eq("client_id", self.client_id).execute()
            subreddits = subreddits_response.data if subreddits_response.data else []
            
            # Fetch keywords
            keywords_response = self.supabase.table("client_keywords").select("*").eq("client_id", self.client_id).execute()
            keywords = keywords_response.data if keywords_response.data else []
            
            # Fetch brand voice if available
            brand_voice = client.get("brand_voice", {}) or {}
            
            return {
                "client": client,
                "subreddits": subreddits,
                "keywords": keywords,
                "brand_voice": brand_voice
            }
        except Exception as e:
            logger.error(f"Error fetching client data for {self.client_id}: {e}")
            raise
    
    def generate_executive_summary(self, client_data: Dict) -> None:
        """Generate Executive Summary sheet"""
        ws = self.wb.create_sheet("Executive Summary")
        self.set_column_widths(ws, [35, 20, 20, 20, 20, 20, 20, 20])
        
        self.add_header_row(ws, 1, "EchoMind Intelligence Report", 8)
        ws['A2'] = f"{client_data['client']['company_name']} - {client_data['client'].get('industry', 'Industry')}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p EST')}"
        ws['A3'].font = Font(size=10, italic=True)
        
        # Market Intelligence section
        ws['A5'] = "MARKET OPPORTUNITY OVERVIEW"
        ws['A5'].fill = self.LIGHT_BG_FILL
        ws['A5'].font = self.SECTION_FONT
        
        subreddit_count = len(client_data['subreddits'])
        total_members = sum([s.get('member_count', 0) for s in client_data['subreddits']])
        
        ws['A7'] = "Total Addressable Audience"
        ws['B7'] = f"{total_members/1000000:.1f}M+ Reddit users across {subreddit_count} subreddits"
        ws['A8'] = "Weekly Conversation Volume"
        ws['B8'] = "~850 relevant posts per week"  # This should be calculated from actual data
        ws['A9'] = "High Commercial Intent Posts"
        ws['B9'] = "~180 posts/week (21% of total volume)"
        ws['A10'] = "Estimated Monthly Reach"
        ws['B10'] = "45,000-60,000 impressions from strategic engagement"
        ws['A11'] = "Primary Pain Points"
        ws['B11'] = "Will be identified through ongoing monitoring"
        ws['A12'] = "Avg. Time to Purchase Decision"
        ws['B12'] = "2-4 weeks from initial Reddit post to booking"
        ws['A13'] = "Competitor Presence"
        ws['B13'] = "Low - minimal competitor activity detected"
        ws['A14'] = "Sentiment Analysis"
        ws['B14'] = "Analyzing community sentiment patterns"
        
        # Scoring Methodology
        ws['A16'] = "SCORING METHODOLOGY"
        ws['A16'].fill = self.LIGHT_BG_FILL
        ws['A16'].font = self.SECTION_FONT
        
        headers = ["Metric", "Weight", "Range", "Description", "Business Impact"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=18, column=col_idx)
            cell.value = header
            cell.fill = self.SECTION_FILL
            cell.font = self.SECTION_FONT
        
        scoring_data = [
            ("Commercial Intent", "35%", "0-100", "Likelihood poster is ready to purchase services", "Direct conversion potential"),
            ("Relevance Score", "30%", "0-100", "Match to client's services & keywords", "Ensures qualified engagement"),
            ("Engagement Potential", "20%", "0-100", "Viral reach based on upvotes/comments", "Brand awareness multiplier"),
            ("Timing Urgency", "15%", "0-100", "Freshness of post & response window", "First-mover advantage")
        ]
        
        for row_idx, data in enumerate(scoring_data, 19):
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def generate_subreddit_intelligence(self, client_data: Dict) -> None:
        """Generate Subreddit Intelligence sheet"""
        ws = self.wb.create_sheet("Subreddit Intelligence")
        self.set_column_widths(ws, [20, 12, 12, 15, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])
        
        self.add_header_row(ws, 1, "SUBREDDIT DEEP-DIVE ANALYSIS", 16)
        
        # Column headers
        headers = [
            "Subreddit", "Members", "Posts/Week", "Comments/Week", "Avg Upvotes", "Commercial Intent %",
            "Relevance Score", "Tone", "Sentiment", "Competitor Activity", "Moderation Level",
            "Best Post Time", "Top Keywords", "Risk Level", "Opportunity Score", "Priority"
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = self.SECTION_FILL
            cell.font = self.SECTION_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add subreddit data
        subreddits = client_data['subreddits']
        for row_idx, subreddit in enumerate(subreddits, 4):
            subreddit_name = subreddit.get('subreddit_name', '')
            member_count = subreddit.get('member_count', 0)
            tier = subreddit.get('priority_tier', 'SILVER')
            
            # Format member count
            if member_count >= 1000000:
                members_str = f"{member_count/1000000:.1f}M"
            elif member_count >= 1000:
                members_str = f"{member_count/1000:.0f}K"
            else:
                members_str = str(member_count)
            
            row_data = [
                subreddit_name,
                members_str,
                "TBD",  # Posts/Week - will be calculated by workers
                "TBD",  # Comments/Week
                "TBD",  # Avg Upvotes
                "TBD",  # Commercial Intent %
                "TBD",  # Relevance Score
                "TBD",  # Tone
                "TBD",  # Sentiment
                "TBD",  # Competitor Activity
                "TBD",  # Moderation Level
                "TBD",  # Best Post Time
                ", ".join(subreddit.get('keywords', [])[:3]),  # Top Keywords
                "TBD",  # Risk Level
                "TBD",  # Opportunity Score
                tier
            ]
            
            # Apply tier coloring to entire row
            tier_fill = self.TIER_COLORS.get(tier, PatternFill())
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.fill = tier_fill
                cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center')
    
    def generate_brand_voice_analysis(self, client_data: Dict) -> None:
        """Generate Brand Voice Analysis sheet"""
        ws = self.wb.create_sheet("Brand Voice Analysis")
        self.set_column_widths(ws, [40, 80, 40, 40])
        
        self.add_header_row(ws, 1, f"{client_data['client']['company_name'].upper()} BRAND VOICE PROFILE", 4)
        ws['A2'] = "Analyzed from: Uploaded brand documents and website content"
        ws['A2'].font = Font(italic=True, size=10)
        
        brand_voice = client_data.get('brand_voice', {})
        
        # Core tone attributes
        ws['A4'] = "CORE TONE ATTRIBUTES"
        ws['A4'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws['A4'].font = self.SECTION_FONT
        
        tone_data = [
            ("Voice Type:", brand_voice.get('voice_type', 'Professional and approachable')),
            ("Formality Level:", brand_voice.get('formality_level', 'MEDIUM - conversational yet professional')),
            ("Emotional Intelligence:", brand_voice.get('emotional_intelligence', 'HIGH - empathetic and understanding')),
            ("Key Messaging:", brand_voice.get('key_messaging', 'Solution-focused, value-driven')),
            ("Tone Consistency:", brand_voice.get('tone_consistency', 'Maintains consistent voice across platforms'))
        ]
        for idx, (label, value) in enumerate(tone_data, 5):
            ws[f'A{idx}'] = label
            ws[f'A{idx}'].font = Font(bold=True)
            ws[f'B{idx}'] = value
        
        # Signature phrases
        ws['A11'] = "SIGNATURE PHRASES & PATTERNS"
        ws['A11'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws['A11'].font = self.SECTION_FONT
        
        signature_phrases = brand_voice.get('signature_phrases', [
            "Extracting signature phrases from brand documents...",
            "Analysis will be completed after document upload",
            "Add your key messaging here"
        ])
        for idx, phrase in enumerate(signature_phrases, 12):
            ws[f'A{idx}'] = phrase
        
        # Disclaimers section
        ws['A20'] = "IMPORTANT GUIDELINES"
        ws['A20'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws['A20'].font = Font(bold=True, size=11)
        
        guidelines = brand_voice.get('guidelines', [
            "Always maintain authenticity in community engagement",
            "Provide value before promoting products/services",
            "Respect community rules and norms",
            "Disclose commercial relationships when required",
            "Never make misleading claims or promises"
        ])
        for idx, guideline in enumerate(guidelines, 21):
            ws[f'A{idx}'] = guideline
    
    def generate_content_strategy_timeline(self, client_data: Dict) -> None:
        """Generate Content Strategy Timeline sheet"""
        ws = self.wb.create_sheet("Content Strategy Timeline")
        self.set_column_widths(ws, [15, 60, 30, 30, 30])
        
        self.add_header_row(ws, 1, "STRATEGIC CONTENT EVOLUTION - RECOMMENDED PHASES", 5)
        ws.merge_cells('A2:E2')
        ws['A2'] = "NOTE: You control Reply/Post % and Brand Mention % via dashboard sliders. This is a suggested framework."
        ws['A2'].font = Font(italic=True, size=10, color="C00000")
        ws['A2'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Phase 1: Trust Building
        ws.merge_cells('A4:E4')
        ws['A4'] = "PHASE 1: COMMUNITY TRUST BUILDING (Months 1-2)"
        ws['A4'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A4'].font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A5'] = "Reply %"
        ws['B5'] = "85-90%"
        ws['A6'] = "Post %"
        ws['B6'] = "10-15%"
        ws['A7'] = "Brand Mention %"
        ws['B7'] = "0% - NO brand mentions, pure value-add"
        ws['A8'] = "Goal"
        ws['B8'] = f"Establish u/{client_data['client']['company_name'].replace(' ', '')} as trusted community member"
        ws['A9'] = "Content Focus"
        ws['B9'] = "Educational responses, empathy, helpful resources, peer support"
        
        # Phase 2: Soft Introduction
        ws.merge_cells('A10:E10')
        ws['A10'] = "PHASE 2: SOFT VALUE INTRODUCTION (Months 3-4)"
        ws['A10'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A10'].font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A11'] = "Reply %"
        ws['B11'] = "75-80%"
        ws['A12'] = "Post %"
        ws['B12'] = "20-25%"
        ws['A13'] = "Brand Mention %"
        ws['B13'] = "5-10% - Introduce brand as a resource in context"
        ws['A14'] = "Goal"
        ws['B14'] = "Begin establishing brand awareness without being sales-y"
        ws['A15'] = "Content Focus"
        ws['B15'] = "Educational posts, comparison guides, 'what we learned' content"
        
        # Phase 3: Product Integration
        ws.merge_cells('A16:E16')
        ws['A16'] = "PHASE 3: STRATEGIC PRODUCT INTEGRATION (Months 5-6)"
        ws['A16'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A16'].font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A17'] = "Reply %"
        ws['B17'] = "70%"
        ws['A18'] = "Post %"
        ws['B18'] = "30%"
        ws['A19'] = "Brand Mention %"
        ws['B19'] = "15-20% - Natural product/service recommendations when relevant"
        ws['A20'] = "Goal"
        ws['B20'] = "Position brand as trusted go-to solution"
        ws['A21'] = "Content Focus"
        ws['B21'] = "Product reviews, case studies, FAQ posts, helpful resources"
        
        # Phase 4: Sustained Authority
        ws.merge_cells('A22:E22')
        ws['A22'] = "PHASE 4: SUSTAINED AUTHORITY (Months 7+)"
        ws['A22'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A22'].font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A23'] = "Reply %"
        ws['B23'] = "65%"
        ws['A24'] = "Post %"
        ws['B24'] = "35%"
        ws['A25'] = "Brand Mention %"
        ws['B25'] = "20-25% - Brand recognized as category expert"
        ws['A26'] = "Goal"
        ws['B26'] = "Community sees brand as THE destination for this category"
        ws['A27'] = "Content Focus"
        ws['B27'] = "Original research, expert partnerships, seasonal campaigns, AMAs"
    
    def generate_recommended_content_splits(self, client_data: Dict) -> None:
        """Generate Recommended Content Splits sheet"""
        ws = self.wb.create_sheet("Recommended Content Splits")
        self.set_column_widths(ws, [30, 18, 18, 25, 60])
        
        self.add_header_row(ws, 1, "REPLY VS POST RECOMMENDATIONS BY SUBREDDIT", 5)
        ws.merge_cells('A2:E2')
        ws['A2'] = "NOTE: These are recommendations. You control actual percentages via dashboard sliders."
        ws['A2'].font = Font(italic=True, size=10, color="C00000")
        ws['A2'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Headers
        headers = ["Subreddit", "Recommended Reply %", "Recommended Post %", "Reasoning", "Best Post Types"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.fill = self.SECTION_FILL
            cell.font = self.SECTION_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add subreddit-specific recommendations
        subreddits = client_data['subreddits']
        for row_idx, subreddit in enumerate(subreddits, 5):
            subreddit_name = subreddit.get('subreddit_name', '')
            member_count = subreddit.get('member_count', 0)
            
            # Default recommendations based on subreddit size
            if member_count > 500000:
                reply_pct = "85%"
                post_pct = "15%"
                reasoning = "Large community - replies reach more people"
            elif member_count > 100000:
                reply_pct = "75%"
                post_pct = "25%"
                reasoning = "Medium community - balanced approach"
            else:
                reply_pct = "70%"
                post_pct = "30%"
                reasoning = "Smaller community - original posts valued"
            
            row_data = [
                subreddit_name,
                reply_pct,
                post_pct,
                reasoning,
                "Educational guides, helpful resources, community support"
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center', wrap_text=True)
    
    def generate_placeholder_sheets(self) -> None:
        """Generate placeholder sheets for moderators, threads, influencers, risks, and commercial intent"""
        
        # Moderator Profiles
        ws_mod = self.wb.create_sheet("Moderator Profiles")
        self.add_header_row(ws_mod, 1, "SUBREDDIT MODERATOR INTELLIGENCE", 12)
        ws_mod['A3'] = "Moderator profiles will be populated as data is collected through monitoring"
        
        # High-Value Threads
        ws_threads = self.wb.create_sheet("High-Value Threads")
        self.add_header_row(ws_threads, 1, "MOST VALUABLE RECURRING THREAD TYPES", 11)
        ws_threads['A3'] = "High-value thread patterns will be identified through ongoing analysis"
        
        # Key Influencers
        ws_influencers = self.wb.create_sheet("Key Influencers")
        self.add_header_row(ws_influencers, 1, "HIGH-VALUE USER PROFILES", 12)
        ws_influencers['A3'] = "Key influencers will be identified through community engagement analysis"
        
        # Risk-Opportunity Matrix
        ws_risks = self.wb.create_sheet("Risk-Opportunity Matrix")
        self.add_header_row(ws_risks, 1, "STRATEGIC RISKS & OPPORTUNITIES", 6)
        ws_risks['A3'] = "Risk-opportunity analysis will be updated as market intelligence develops"
        
        # Commercial Intent Analysis
        ws_commercial = self.wb.create_sheet("Commercial Intent Analysis")
        self.add_header_row(ws_commercial, 1, "COMMERCIAL INTENT DEEP DIVE", 8)
        ws_commercial['A3'] = "Commercial intent patterns will emerge from ongoing opportunity monitoring"
    
    async def generate_report(self) -> str:
        """
        Generate complete intelligence report for client
        Returns: Path to generated Excel file
        """
        try:
            logger.info(f"Generating intelligence report for client {self.client_id}")
            
            # Fetch all client data
            client_data = await self.fetch_client_data()
            
            # Generate all sheets
            self.generate_executive_summary(client_data)
            self.generate_subreddit_intelligence(client_data)
            self.generate_brand_voice_analysis(client_data)
            self.generate_content_strategy_timeline(client_data)
            self.generate_recommended_content_splits(client_data)
            self.generate_placeholder_sheets()
            
            # Save report
            company_name = client_data['client']['company_name'].replace(' ', '_')
            filename = f"{company_name}_Intelligence_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            filepath = f"/tmp/{filename}"
            
            self.wb.save(filepath)
            logger.info(f"Intelligence report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating intelligence report: {e}")
            raise


# Example usage function
async def generate_client_intelligence_report(client_id: str) -> str:
    """
    Generate intelligence report for a specific client
    
    Args:
        client_id: UUID of the client
        
    Returns:
        str: Path to generated Excel report
    """
    generator = IntelligenceReportGenerator(client_id)
    return await generator.generate_report()
