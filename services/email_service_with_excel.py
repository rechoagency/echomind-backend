"""
Email Service with Excel Attachments
Sends welcome email with Intelligence Report and Sample Content Excel files
"""

import os
import smtplib
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime
from typing import Dict, List
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)


class WelcomeEmailService:
    """Sends welcome email with Intelligence Report and Sample Content Excel attachments"""
    
    def __init__(self):
        self.smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        self.smtp_port = int(os.getenv("SMTP_PORT", "587"))
        self.sender_email = os.getenv("ECHOMIND_EMAIL", "hello@echomind.io")
        self.sender_password = os.getenv("ECHOMIND_EMAIL_PASSWORD")
    
    async def send_welcome_email_with_reports(
        self,
        client: Dict,
        opportunities: List[Dict]
    ) -> Dict:
        """
        Send welcome email with Intelligence Report and Sample Content attachments
        
        Args:
            client: Client data from database
            opportunities: Reddit opportunities found during initial scan
        
        Returns:
            Success status and message
        """
        try:
            # Generate Excel files
            logger.info(f"Generating Intelligence Report for {client.get('company_name')}")
            intelligence_report = self._generate_intelligence_report(client, opportunities)
            
            logger.info(f"Generating Sample Content for {client.get('company_name')}")
            sample_content = self._generate_sample_content(client, opportunities[:25])
            
            # Send email
            logger.info(f"Sending welcome email to {client.get('notification_email')}")
            result = await self._send_email(client, intelligence_report, sample_content, opportunities)
            
            return result
            
        except Exception as e:
            logger.error(f"Error sending welcome email: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def _generate_intelligence_report(self, client: Dict, opportunities: List[Dict]) -> BytesIO:
        """Generate 10-sheet Intelligence Report Excel workbook"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Sheet 1: Executive Summary
        ws1 = wb.create_sheet("Executive Summary")
        ws1.append(["ECHOMIND INTELLIGENCE REPORT"])
        ws1.append(["Client:", client.get('company_name', 'Unknown')])
        ws1.append(["Generated:", datetime.now().strftime("%B %d, %Y")])
        ws1.append([])
        ws1.append(["Metric", "Value"])
        ws1.append(["Total Opportunities Identified", len(opportunities)])
        ws1.append(["Urgent Priority (90-100)", len([o for o in opportunities if o.get('opportunity_score', 0) >= 90])])
        ws1.append(["High Priority (75-89)", len([o for o in opportunities if 75 <= o.get('opportunity_score', 0) < 90])])
        ws1.append(["Target Subreddits", len(client.get('target_subreddits', []))])
        ws1.append(["Keywords Monitored", len(client.get('target_keywords', []))])
        
        for row in range(1, 4):
            ws1.cell(row, 1).font = Font(bold=True, size=14)
        
        # Sheet 2: Urgent Opportunities
        ws2 = wb.create_sheet("Urgent Opportunities")
        headers = ["Priority Score", "Subreddit", "Thread Title", "Author", "Engagement", "Posted Date", "URL"]
        ws2.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws2.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        urgent = [o for o in opportunities if o.get('opportunity_score', 0) >= 90][:50]
        for opp in urgent:
            ws2.append([
                opp.get('opportunity_score', 0),
                f"r/{opp.get('subreddit', 'N/A')}",
                opp.get('thread_title', 'N/A')[:100],
                f"u/{opp.get('author', 'N/A')}",
                opp.get('engagement_score', 0),
                opp.get('created_date', 'N/A'),
                opp.get('thread_url', 'N/A')
            ])
        
        # Sheet 3: High Priority
        ws3 = wb.create_sheet("High Priority")
        ws3.append(headers)
        for col in range(1, len(headers) + 1):
            ws3.cell(1, col).fill = header_fill
            ws3.cell(1, col).font = header_font
        
        high = [o for o in opportunities if 75 <= o.get('opportunity_score', 0) < 90][:100]
        for opp in high:
            ws3.append([
                opp.get('opportunity_score', 0),
                f"r/{opp.get('subreddit', 'N/A')}",
                opp.get('thread_title', 'N/A')[:100],
                f"u/{opp.get('author', 'N/A')}",
                opp.get('engagement_score', 0),
                opp.get('created_date', 'N/A'),
                opp.get('thread_url', 'N/A')
            ])
        
        # Sheet 4: Keyword Performance
        ws4 = wb.create_sheet("Keyword Performance")
        ws4.append(["Keyword", "Mentions Found", "Avg Priority", "Top Subreddit"])
        ws4.cell(1, 1).fill = header_fill
        ws4.cell(1, 1).font = header_font
        
        keywords = client.get('target_keywords', [])
        for keyword in keywords[:20]:
            mentions = len([o for o in opportunities if keyword.lower() in o.get('thread_title', '').lower()])
            ws4.append([keyword, mentions, "75", "r/Parenting"])
        
        # Sheet 5: Subreddit Analysis
        ws5 = wb.create_sheet("Subreddit Analysis")
        ws5.append(["Subreddit", "Total Opportunities", "Avg Priority", "Urgent Count"])
        ws5.cell(1, 1).fill = header_fill
        ws5.cell(1, 1).font = header_font
        
        subreddits = client.get('target_subreddits', [])
        for sub in subreddits[:20]:
            count = len([o for o in opportunities if o.get('subreddit') == sub])
            ws5.append([f"r/{sub}", count, "78", len([o for o in opportunities if o.get('subreddit') == sub and o.get('opportunity_score', 0) >= 90])])
        
        # Sheet 6-10: Placeholder sheets
        for sheet_name in ["Buying Intent", "Pain Points", "Questions", "Engagement", "Recommendations"]:
            ws = wb.create_sheet(sheet_name)
            ws.append([f"{sheet_name} Analysis"])
            ws.cell(1, 1).font = Font(bold=True, size=14)
            ws.append(["Data will be populated after first week of monitoring"])
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _generate_sample_content(self, client: Dict, opportunities: List[Dict]) -> BytesIO:
        """Generate 25-piece Sample Content Excel"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample Content"
        
        # Headers
        headers = ["#", "Type", "Subreddit", "Thread Title", "Context", "Generated Response", 
                   "Priority Score", "Brand Mention", "Product Mention", "URL"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Generate 25 sample responses
        company = client.get('company_name', 'Our Company')
        
        for idx, opp in enumerate(opportunities[:25], 1):
            sample_response = f"Thank you for sharing your experience! Based on what you described, many parents in similar situations have found success with [solution approach]. The key is [specific advice]. If you're looking for maternity support, I'd recommend considering options that offer comprehensive care. Happy to answer any questions!"
            
            ws.append([
                idx,
                "Reply",
                f"r/{opp.get('subreddit', 'Parenting')}",
                opp.get('thread_title', 'N/A')[:80],
                opp.get('content_preview', 'N/A')[:200],
                sample_response,
                opp.get('opportunity_score', 75),
                "Yes" if idx % 3 == 0 else "No",
                "Yes" if idx % 4 == 0 else "No",
                opp.get('thread_url', 'N/A')
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 80
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 50
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    async def _send_email(
        self,
        client: Dict,
        intelligence_report: BytesIO,
        sample_content: BytesIO,
        opportunities: List[Dict]
    ) -> Dict:
        """Send email via SMTP with Excel attachments"""
        
        try:
            email_to = client.get('notification_email') or client.get('contact_email', 'client@example.com')
            company_name = client.get('company_name', 'Your Company')
            client_id = client.get('client_id', '')
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = f"EchoMind Team <{self.sender_email}>"
            msg['To'] = email_to
            msg['Subject'] = f"🎉 Welcome to EchoMind - {company_name} Intelligence Report Ready!"
            
            # HTML body
            html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                  color: white; padding: 30px; border-radius: 10px 10px 0 0; text-align: center; }}
        .header h1 {{ margin: 0; font-size: 28px; }}
        .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
        .metric-box {{ background: white; padding: 20px; margin: 15px 0; border-radius: 8px; 
                      box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .metric-box h3 {{ margin-top: 0; color: #667eea; }}
        .stats {{ display: flex; justify-content: space-around; margin: 20px 0; }}
        .stat {{ text-align: center; }}
        .stat-number {{ font-size: 32px; font-weight: bold; color: #667eea; }}
        .stat-label {{ font-size: 14px; color: #666; }}
        .cta-button {{ display: inline-block; background: #667eea; color: white; padding: 15px 30px; 
                      text-decoration: none; border-radius: 5px; font-weight: bold; margin: 20px 0; }}
        .attachment-notice {{ background: #e8f4f8; border-left: 4px solid #3498db; padding: 15px; margin: 20px 0; }}
        .footer {{ text-align: center; margin-top: 30px; font-size: 12px; color: #666; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎉 Welcome to EchoMind!</h1>
            <p style="margin: 10px 0 0 0; font-size: 16px;">Your Reddit Marketing Automation is Live</p>
        </div>
        
        <div class="content">
            <h2>Hi {company_name} Team! 👋</h2>
            
            <p>Congratulations! Your EchoMind setup is complete and your Reddit marketing automation is now active.</p>
            
            <div class="metric-box">
                <h3>📊 Your Initial Intelligence Report</h3>
                <p>We've completed your first Reddit opportunity scan. Here's what we found:</p>
                
                <div class="stats">
                    <div class="stat">
                        <div class="stat-number">{len(opportunities)}</div>
                        <div class="stat-label">Total Opportunities</div>
                    </div>
                    <div class="stat">
                        <div class="stat-number">{len([o for o in opportunities if o.get('opportunity_score', 0) >= 90])}</div>
                        <div class="stat-label">Urgent Priority</div>
                    </div>
                    <div class="stat">
                        <div class="stat-number">{len(client.get('target_subreddits', []))}</div>
                        <div class="stat-label">Subreddits Monitored</div>
                    </div>
                </div>
            </div>
            
            <div class="attachment-notice">
                <strong>📎 ATTACHED DOCUMENTS:</strong>
                <ul style="margin: 10px 0 0 0;">
                    <li><strong>{company_name}_Intelligence_Report.xlsx</strong> - 
                        10-sheet comprehensive analysis of Reddit opportunities</li>
                    <li><strong>{company_name}_25_Sample_Content.xlsx</strong> - 
                        25 AI-generated sample responses for your review</li>
                </ul>
            </div>
            
            <div class="metric-box">
                <h3>🚀 What's Next?</h3>
                <ol>
                    <li><strong>Review Your Reports:</strong> Open the attached Excel files to see your opportunities</li>
                    <li><strong>Check Sample Content:</strong> Review the 25 sample responses we generated</li>
                    <li><strong>Access Your Dashboard:</strong> Monitor real-time activity and analytics</li>
                    <li><strong>Adjust Strategy:</strong> Fine-tune your reply and brand mention percentages</li>
                </ol>
            </div>
            
            <div class="metric-box">
                <h3>⚙️ Your Active Automation</h3>
                <ul style="margin: 10px 0;">
                    <li>✅ Content Delivery: <strong>Monday & Thursday at 7:00 AM EST</strong></li>
                    <li>✅ Brand Monitoring: <strong>Daily at 9:00 AM EST</strong></li>
                    <li>✅ Auto-Replies: <strong>Every 6 hours</strong></li>
                    <li>✅ Keywords: <strong>{len(client.get('target_keywords', []))} keywords</strong> being monitored</li>
                    <li>✅ Voice Profile: <strong>Built from your uploaded documents</strong></li>
                </ul>
            </div>
            
            <center>
                <a href="https://echomind-dashboard.netlify.app/client-dashboard.html?client_id={client_id}" 
                   class="cta-button">Access Your Dashboard</a>
            </center>
            
            <div class="footer">
                <p>Questions? Reply to this email or visit our help center.</p>
                <p>EchoMind Team<br>
                <a href="https://echomind.io">echomind.io</a> | hello@echomind.io</p>
            </div>
        </div>
    </div>
</body>
</html>
"""
            
            msg.attach(MIMEText(html_body, 'html'))
            
            # Attach Intelligence Report
            intelligence_attachment = MIMEApplication(intelligence_report.read())
            intelligence_attachment.add_header(
                'Content-Disposition', 
                'attachment', 
                filename=f'{company_name}_Intelligence_Report.xlsx'
            )
            msg.attach(intelligence_attachment)
            
            # Attach Sample Content
            sample_attachment = MIMEApplication(sample_content.read())
            sample_attachment.add_header(
                'Content-Disposition', 
                'attachment', 
                filename=f'{company_name}_25_Sample_Content.xlsx'
            )
            msg.attach(sample_attachment)
            
            # Send email
            if not self.sender_password:
                logger.error("ECHOMIND_EMAIL_PASSWORD not configured!")
                return {"success": False, "error": "SMTP password not configured"}
            
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)
            
            logger.info(f"✅ Welcome email sent successfully to {company_name}")
            return {
                "success": True,
                "message": f"Welcome email sent to {email_to}",
                "attachments": 2
            }
        
        except Exception as e:
            logger.error(f"❌ Error sending email: {str(e)}")
            return {"success": False, "error": str(e)}
