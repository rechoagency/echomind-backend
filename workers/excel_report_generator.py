"""
Excel Weekly Report Generator v2.0

Replaces HTML email reports with Excel workbooks.
Generates 25-piece content queues every Monday & Thursday at 7am EST.

NEW FORMAT (30-31 columns, A through AD):
- Voice matching columns (formality, tone, similarity proof)
- Anti-AI detection tracking
- Full opportunity and content metadata
- Product matchback and knowledge base usage
"""

import os
import asyncio
import logging
from typing import Dict, List, Any
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from database import get_supabase_client
from workers.enhanced_content_generation_worker import generate_enhanced_content

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """Generates Excel weekly content reports"""
    
    def __init__(self):
        self.supabase = get_supabase_client()
    
    async def generate_weekly_report(
        self,
        client_id: str,
        brand_mention_percentage: float = 0.0,
        num_opportunities: int = 25
    ) -> str:
        """
        Generate Excel report with content queue
        
        Args:
            client_id: UUID of client
            brand_mention_percentage: 0-100 control
            num_opportunities: Number of opportunities to include (default 25)
            
        Returns:
            File path to generated Excel report
        """
        try:
            logger.info(f"Generating Excel report for client {client_id}")
            
            # Fetch client info
            client_response = self.supabase.table("clients").select("*").eq("client_id", client_id).execute()
            if not client_response.data:
                raise ValueError(f"Client {client_id} not found")
            
            client = client_response.data[0]
            company_name = client.get('company_name', 'Client')
            
            # Fetch top opportunities from last 3 days
            three_days_ago = (datetime.utcnow() - timedelta(days=3)).isoformat()
            opportunities_response = self.supabase.table("opportunities")\
                .select("*")\
                .eq("client_id", client_id)\
                .gte("created_at", three_days_ago)\
                .order("combined_score", desc=True)\
                .limit(num_opportunities)\
                .execute()
            
            opportunities = opportunities_response.data
            
            if not opportunities:
                logger.warning(f"No opportunities found for client {client_id}")
                return None
            
            logger.info(f"Found {len(opportunities)} opportunities for report")
            
            # Generate content for each opportunity
            content_queue = []
            for rank, opp in enumerate(opportunities, 1):
                try:
                    content_data = await generate_enhanced_content(
                        opp['id'],
                        client_id,
                        brand_mention_percentage
                    )
                    
                    content_queue.append({
                        "opportunity": opp,
                        "content": content_data,
                        "rank": rank
                    })
                    
                except Exception as e:
                    logger.error(f"Failed to generate content for opportunity {opp['id']}: {e}")
                    continue
            
            # Create Excel workbook
            filepath = await self._create_excel_report(
                client,
                content_queue,
                brand_mention_percentage
            )
            
            logger.info(f"Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise
    
    async def _create_excel_report(
        self,
        client: Dict,
        content_queue: List[Dict],
        brand_mention_percentage: float
    ) -> str:
        """Create Excel workbook with content queue"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Content Queue"
        
        # Define styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Define headers (30 columns: A through AD)
        # User-specified format for anti-AI voice-matched content
        headers = [
            "Opportunity ID",           # A
            "Subreddit",                # B
            "Thread URL",               # C
            "Thread Title",             # D
            "Original Post",            # E
            "Author Username",          # F
            "Date Posted",              # G
            "Date Found",               # H
            "Matched Keywords",         # I
            "Urgency",                  # J
            "Content Type",             # K
            "Generated Reply",          # L (THE CONTENT - ready to copy/paste)
            "Word Count",               # M
            "Voice Formality Score",    # N
            "Voice Tone",               # O
            "Voice Similarity Proof",   # P
            "Typos Injected",           # Q
            "AI Violations Detected",   # R
            "Regeneration Attempts",    # S
            "Brand Mentioned",          # T
            "Product Mentioned",        # U
            "Product Similarity",       # V
            "Knowledge Base Used",      # W
            "Knowledge Excerpts",       # X
            "Assigned Profile",         # Y
            "Profile Karma",            # Z
            "Combined Score",           # AA
            "Content Status",           # AB
            "Notes",                    # AC
            "Posting Account",          # AD
        ]
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths (30 columns)
        column_widths = [
            15,   # A: Opportunity ID
            14,   # B: Subreddit
            50,   # C: Thread URL
            50,   # D: Thread Title
            60,   # E: Original Post
            16,   # F: Author Username
            18,   # G: Date Posted
            18,   # H: Date Found
            25,   # I: Matched Keywords
            12,   # J: Urgency
            14,   # K: Content Type
            120,  # L: Generated Reply (wide for copy/paste)
            10,   # M: Word Count
            14,   # N: Voice Formality Score
            20,   # O: Voice Tone
            50,   # P: Voice Similarity Proof
            10,   # Q: Typos Injected
            14,   # R: AI Violations Detected
            14,   # S: Regeneration Attempts
            12,   # T: Brand Mentioned
            12,   # U: Product Mentioned
            12,   # V: Product Similarity
            14,   # W: Knowledge Base Used
            50,   # X: Knowledge Excerpts
            18,   # Y: Assigned Profile
            10,   # Z: Profile Karma
            12,   # AA: Combined Score
            14,   # AB: Content Status
            40,   # AC: Notes
            18,   # AD: Posting Account
        ]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        
        # Add content rows
        for row_idx, item in enumerate(content_queue, 2):
            opp = item['opportunity']
            content = item['content']
            rank = item['rank']

            # Determine urgency
            timing = self._get_timing_category(opp.get('created_at') or opp.get('date_found'))
            urgency = self._get_urgency_emoji(opp.get('combined_score', 0), timing)

            # Extract matched keywords (may be JSON string or list)
            matched_keywords = opp.get('matched_keywords', '')
            if isinstance(matched_keywords, list):
                matched_keywords = ', '.join(matched_keywords)
            elif matched_keywords and matched_keywords.startswith('['):
                try:
                    import json
                    matched_keywords = ', '.join(json.loads(matched_keywords))
                except:
                    pass

            # Build row data (30 columns: A through AD)
            row_data = [
                opp.get('opportunity_id', opp.get('id', ''))[:15],  # A: Opportunity ID
                opp.get('subreddit', ''),                           # B: Subreddit
                opp.get('thread_url', ''),                          # C: Thread URL
                opp.get('thread_title', ''),                        # D: Thread Title
                (opp.get('original_post_text', '') or '')[:500],    # E: Original Post
                opp.get('author_username', ''),                     # F: Author Username
                opp.get('date_posted', ''),                         # G: Date Posted
                opp.get('date_found', ''),                          # H: Date Found
                matched_keywords,                                    # I: Matched Keywords
                urgency,                                             # J: Urgency
                content.get('type', 'REPLY').upper(),               # K: Content Type
                content.get('text', content.get('content', '')),    # L: Generated Reply
                content.get('actual_word_count', len((content.get('text', '') or '').split())),  # M: Word Count
                round(content.get('formality_score', 0.5), 2),      # N: Voice Formality Score
                content.get('tone', 'conversational'),               # O: Voice Tone
                content.get('voice_similarity_proof', ''),           # P: Voice Similarity Proof
                content.get('typos_injected', 0),                    # Q: Typos Injected
                content.get('ai_violations_detected', 0),            # R: AI Violations Detected
                content.get('regeneration_attempts', 1),             # S: Regeneration Attempts
                'Yes' if content.get('brand_mentioned') else 'No',   # T: Brand Mentioned
                'Yes' if content.get('product_mentioned') else 'No', # U: Product Mentioned
                round(opp.get('product_similarity', 0), 2),          # V: Product Similarity
                content.get('knowledge_insights_used', 0),           # W: Knowledge Base Used
                '; '.join(content.get('knowledge_excerpts', []))[:200],  # X: Knowledge Excerpts
                content.get('assigned_profile', ''),                 # Y: Assigned Profile
                content.get('profile_karma', 0),                     # Z: Profile Karma
                round(opp.get('combined_score', opp.get('overall_priority', 0)), 2),  # AA: Combined Score
                "Ready to Post",                                     # AB: Content Status
                self._build_notes(opp, content, brand_mention_percentage),  # AC: Notes
                content.get('assigned_profile', ''),                 # AD: Posting Account
            ]

            # Write row
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border

                # Special formatting for specific columns
                # L (12): Generated Reply, E (5): Original Post, P (16): Voice Proof, X (24): Knowledge
                if col_idx in [5, 12, 16, 24, 29]:  # Wrap text columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                elif col_idx == 10:  # J: Urgency colors
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if "🔴" in str(value):
                        cell.font = Font(color="FF0000", bold=True)
                    elif "🟡" in str(value):
                        cell.font = Font(color="FFA500", bold=True)
                    elif "🟢" in str(value):
                        cell.font = Font(color="008000", bold=True)
                else:
                    cell.alignment = Alignment(vertical='center')
        
        # Save file
        company_name_clean = client.get('company_name', 'Client').replace(' ', '_')
        timestamp = datetime.utcnow().strftime("%Y%m%d")
        filename = f"{company_name_clean}_Weekly_Content_{timestamp}.xlsx"
        filepath = f"/tmp/{filename}"
        
        wb.save(filepath)
        return filepath
    
    def _get_priority_tier(self, score: float) -> str:
        """Determine priority tier from score"""
        if score >= 0.8:
            return "TIER_1 (HIGHEST)"
        elif score >= 0.65:
            return "TIER_2 (HIGH)"
        elif score >= 0.5:
            return "TIER_3 (MEDIUM)"
        else:
            return "TIER_4 (LOW)"
    
    def _get_timing_category(self, created_at: str) -> str:
        """Determine timing urgency"""
        if not created_at:
            return "UNKNOWN"
        
        created = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        age_hours = (datetime.utcnow() - created.replace(tzinfo=None)).total_seconds() / 3600
        
        if age_hours <= 4:
            return "IMMEDIATE"
        elif age_hours <= 48:
            return "WITHIN_48H"
        elif age_hours <= 96:
            return "WITHIN_4DAYS"
        else:
            return "EXPIRED"
    
    def _get_urgency_emoji(self, score: float, timing: str) -> str:
        """Get urgency emoji based on score and timing"""
        if score >= 0.8 and timing in ["IMMEDIATE", "WITHIN_48H"]:
            return "🔴 URGENT"
        elif score >= 0.65:
            return "🟡 HIGH"
        else:
            return "🟢 MEDIUM"
    
    def _build_notes(self, opp: Dict, content: Dict, brand_percentage: float) -> str:
        """Build notes column with metadata for v2.0 format"""
        notes = []

        # Priority indicator
        combined_score = opp.get('combined_score', opp.get('overall_priority', 0))
        if combined_score >= 80 or (combined_score >= 0.8 and combined_score <= 1):
            notes.append("PLATINUM OPPORTUNITY")
        elif combined_score >= 70 or (combined_score >= 0.7 and combined_score <= 1):
            notes.append("HIGH-VALUE")

        # Voice matching info
        formality = content.get('formality_score', 0.5)
        if formality < 0.3:
            notes.append("Voice: Very Casual")
        elif formality < 0.5:
            notes.append("Voice: Casual")
        elif formality < 0.7:
            notes.append("Voice: Conversational")
        else:
            notes.append("Voice: Semi-formal")

        # AI detection info
        ai_violations = content.get('ai_violations_detected', 0)
        if ai_violations == 0:
            notes.append("AI-Clean")
        else:
            notes.append(f"AI-Flagged ({ai_violations} patterns)")

        # Typo injection info
        typos = content.get('typos_injected', 0)
        if typos > 0:
            notes.append(f"{typos} typo(s) added")

        # Knowledge base usage
        kb_used = content.get('knowledge_insights_used', 0)
        if kb_used > 0:
            notes.append(f"KB: {kb_used} insights")

        return " | ".join(notes)


async def generate_weekly_excel_report(
    client_id: str,
    brand_mention_percentage: float = 0.0
) -> str:
    """
    Generate weekly Excel report for client
    
    Args:
        client_id: UUID of client
        brand_mention_percentage: 0-100
        
    Returns:
        Filepath to generated Excel report
    """
    generator = ExcelReportGenerator()
    return await generator.generate_weekly_report(client_id, brand_mention_percentage)
